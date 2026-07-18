from rest_framework import viewsets, status, serializers
from rest_framework.decorators import action
from rest_framework.response import Response
from decimal import Decimal
from django.core.exceptions import ValidationError as DjangoValidationError
from user.permissions import IsEmployee
from .models import Taminotchi, SupplierOrder, SupplierOrderReturn
from .serializers import TaminotchiSerializer, SupplierOrderSerializer, SupplierOrderReturnSerializer

class TaminotchiViewSet(viewsets.ModelViewSet):
    serializer_class = TaminotchiSerializer
    permission_classes = [IsEmployee]
    search_fields = ['id', 'nomi', 'telefon_raqam']
    ordering_fields = ['nomi']

    def get_queryset(self):
        user = self.request.user
        queryset = Taminotchi.objects.all().order_by('-yaratilgan_vaqt')
        if user.is_superuser:
            return queryset
        if user.is_authenticated and hasattr(user, 'xodim') and user.xodim.biznes:
            return queryset.filter(biznes=user.xodim.biznes)
        return queryset.none()

    def perform_create(self, serializer):
        biznes = None
        if self.request.user and hasattr(self.request.user, 'xodim'):
            biznes = self.request.user.xodim.biznes
        serializer.save(biznes=biznes)

import django_filters
from django.db import models

class SupplierOrderFilter(django_filters.FilterSet):
    to_lov_status = django_filters.CharFilter(method='filter_tolov_status')
    tolov_status = django_filters.CharFilter(method='filter_tolov_status')

    class Meta:
        model = SupplierOrder
        fields = ['holat', 'taminotchi', 'dokon', 'to_lov_status', 'tolov_status']

    def filter_tolov_status(self, queryset, name, value):
        if value == 'tolanmagan':
            return queryset.filter(tolangan_summa=0)
        elif value == 'qisman_tolangan':
            return queryset.filter(tolangan_summa__gt=0, tolangan_summa__lt=models.F('umumiy_summa'))
        elif value == 'tolangan':
            return queryset.filter(tolangan_summa=models.F('umumiy_summa'), umumiy_summa__gt=0)
        return queryset


class SupplierOrderViewSet(viewsets.ModelViewSet):
    serializer_class = SupplierOrderSerializer
    permission_classes = [IsEmployee]
    filterset_class = SupplierOrderFilter
    search_fields = ['id', 'nomi', 'taminotchi__nomi']
    ordering_fields = ['umumiy_summa', 'nasiya_summa', 'yaratilgan_vaqt']

    def get_permissions(self):
        from rest_framework.permissions import AllowAny
        if self.action == 'template':
            return [AllowAny()]
        return super().get_permissions()

    @action(detail=False, methods=['get'])
    def template(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Buyurtma Shablon"

        # Headers
        headers = [
            "Nomi", "Shtrix-kod", "Buyurtmaga", "Kelish narxi", "Ustama %", "Sotuv narxi", "Ulgurji narx"
        ]
        ws.append(headers)

        # Style headers
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 5, 12)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=buyurtma_template.xlsx'
        wb.save(response)

        return response

    def list(self, request, *args, **kwargs):
        if request.query_params.get('export') == 'excel':
            queryset = self.filter_queryset(self.get_queryset())
            headers = ["ID", "Nomi", "Do'kon", "Holat", "To'lov status", "Miqdori", "Buyurtma summasi", "Yaratildi", "Qabul qilindi"]
            rows = []
            for item in queryset:
                serializer = self.get_serializer(item)
                rows.append([
                    item.id,
                    item.nomi,
                    item.dokon.nomi if item.dokon else "",
                    item.get_holat_display() if hasattr(item, 'get_holat_display') else item.holat,
                    serializer.data.get('tolov_status', ''),
                    item.elementlar.aggregate(total=models.Sum('miqdori'))['total'] or 0,
                    str(item.umumiy_summa),
                    item.yaratilgan_vaqt.strftime("%d.%m.%Y %H:%M:%S") if item.yaratilgan_vaqt else "",
                    item.haqiqiy_qabul_sana.strftime("%d.%m.%Y %H:%M:%S") if item.haqiqiy_qabul_sana else ""
                ])
            from products.views import generate_excel_response
            return generate_excel_response("buyurtmalar", headers, rows)
        return super().list(request, *args, **kwargs)

    def get_queryset(self):
        user = self.request.user
        queryset = SupplierOrder.objects.all().order_by('-yaratilgan_vaqt')
        if user.is_superuser:
            return queryset
        if user.is_authenticated and hasattr(user, 'xodim') and user.xodim.biznes:
            return queryset.filter(biznes=user.xodim.biznes)
        return queryset.none()

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        if instance.holat != 'qoralama':
            return Response(
                {"detail": "Faqat qoralama holatidagi buyurtmalarni o'zgartirish mumkin."},
                status=status.HTTP_400_BAD_REQUEST
            )
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if instance.holat != 'qoralama':
            return Response(
                {"detail": "Faqat qoralama holatidagi buyurtmalarni o'chirish mumkin."},
                status=status.HTTP_400_BAD_REQUEST
            )
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['post'])
    def confirm(self, request, pk=None):
        order_obj = self.get_object()
        try:
            order_obj.rasmiylashtirish()
        except DjangoValidationError as e:
            raise serializers.ValidationError({'detail': str(e)})
        return Response({
            'status': "Buyurtma muvaffaqiyatli rasmiylashtirildi.",
            'holat': order_obj.holat
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'])
    def pay(self, request, pk=None):
        order_obj = self.get_object()
        amount = request.data.get('amount')
        tolov_turi = request.data.get('tolov_turi')
        
        if not amount or not tolov_turi:
            return Response({"detail": "amount va tolov_turi maydonlari kiritilishi shart."}, status=status.HTTP_400_BAD_REQUEST)
            
        try:
            amount_decimal = Decimal(str(amount))
        except Exception:
            return Response({"detail": "To'lov summasi noto'g'ri ko'rinishda."}, status=status.HTTP_400_BAD_REQUEST)

        xodim = request.user.xodim if hasattr(request.user, 'xodim') else None
        try:
            order_obj.add_payment(amount_decimal, tolov_turi, xodim)
        except DjangoValidationError as e:
            raise serializers.ValidationError({'detail': str(e)})

        return Response({
            'status': "To'lov muvaffaqiyatli amalga oshirildi.",
            'tolangan_summa': str(order_obj.tolangan_summa),
            'nasiya_summa': str(order_obj.nasiya_summa)
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['get'])
    def price_differences(self, request, pk=None):
        order_obj = self.get_object()
        diffs = order_obj.get_price_differences()
        return Response(diffs, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'])
    def receive(self, request, pk=None):
        order_obj = self.get_object()
        apply_new_prices = request.data.get('apply_new_prices', False)
        
        if not isinstance(apply_new_prices, bool):
            apply_new_prices = str(apply_new_prices).lower() == 'true'

        xodim = request.user.xodim if hasattr(request.user, 'xodim') else None
        try:
            order_obj.qabul_qilish(apply_new_prices, xodim)
        except DjangoValidationError as e:
            raise serializers.ValidationError({'detail': str(e)})

        return Response({
            'status': "Mahsulotlar muvaffaqiyatli qabul qilindi va do'kon qoldig'iga qo'shildi.",
            'holat': order_obj.holat
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        order_obj = self.get_object()
        try:
            order_obj.bekor_qilish()
        except DjangoValidationError as e:
            raise serializers.ValidationError({'detail': str(e)})
        return Response({
            'status': "Buyurtma bekor qilindi.",
            'holat': order_obj.holat
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'])
    def template(self, request):
        import openpyxl
        from django.http import HttpResponse
        from decimal import Decimal
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Shablon"
        
        headers = ["Nomi", "Shtrix-kod", "Buyurtmaga", "Kelish narxi", "Ustama %", "Sotuv narxi", "Ulgurji narx"]
        ws.append(headers)
        ws.append(["Armatura", "9948493123", "500", "40000", "50", "60000", "64000"])
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=buyurtma_shablon.xlsx'
        wb.save(response)
        return response

class SupplierOrderReturnViewSet(viewsets.ModelViewSet):
    serializer_class = SupplierOrderReturnSerializer
    permission_classes = [IsEmployee]
    filterset_fields = ['holat', 'taminotchi', 'dokon']
    search_fields = ['order__nomi', 'taminotchi__nomi']
    ordering_fields = ['qaytarish_summasi', 'yaratilgan_vaqt']

    def get_queryset(self):
        user = self.request.user
        queryset = SupplierOrderReturn.objects.all().order_by('-yaratilgan_vaqt')
        if user.is_superuser:
            return queryset
        if user.is_authenticated and hasattr(user, 'xodim') and user.xodim.biznes:
            return queryset.filter(biznes=user.xodim.biznes)
        return queryset.none()

    @action(detail=True, methods=['post'])
    def confirm(self, request, pk=None):
        return_obj = self.get_object()
        try:
            return_obj.execute_return()
        except DjangoValidationError as e:
            raise serializers.ValidationError({'detail': str(e)})
        return Response({
            'status': "Tovarlar yetkazib beruvchiga muvaffaqiyatli qaytarildi va balans yangilandi.",
            'holat': return_obj.holat
        }, status=status.HTTP_200_OK)
