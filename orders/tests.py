from django.urls import reverse
from rest_framework import status
from rest_framework.test import APITestCase
from django.core.exceptions import ValidationError
from decimal import Decimal
from io import BytesIO
import openpyxl
from rest_framework.authtoken.models import Token
from user.models import Biznes, Xodim, Tarif
from products.models import Mahsulot, Dokon, DokonQoldiq, MahsulotShtrixKod
from orders.models import Taminotchi, SupplierOrder, SupplierOrderItem, SupplierOrderPayment, SupplierOrderReturn, SupplierOrderReturnItem

class SupplierOrdersAPITestCase(APITestCase):
    def setUp(self):
        # 1. Create subscriptions and businesses
        self.tarif = Tarif.objects.create(nomi="Standard", xodim_limiti=5, dokon_limiti=5, mahsulot_limiti=50)
        self.biznes1 = Biznes.objects.create(nomi="Biznes A", egasi_ism="Owner A", tarif=self.tarif)
        self.biznes2 = Biznes.objects.create(nomi="Biznes B", egasi_ism="Owner B", tarif=self.tarif)

        # 2. Create stores
        self.dokon1 = Dokon.objects.create(biznes=self.biznes1, nomi="Do'kon A")
        self.dokon2 = Dokon.objects.create(biznes=self.biznes2, nomi="Do'kon B")

        # 3. Create employees
        self.admin1 = Xodim.objects.create(
            biznes=self.biznes1, ism="Admin", familiya="A", telefon_raqam="+998901000001",
            parol="SecurePass1!", rol="admin", jinsi="erkak"
        )
        self.token1 = Token.objects.create(user=self.admin1.user).key

        self.admin2 = Xodim.objects.create(
            biznes=self.biznes2, ism="Admin", familiya="B", telefon_raqam="+998901000002",
            parol="SecurePass2!", rol="admin", jinsi="erkak"
        )
        self.token2 = Token.objects.create(user=self.admin2.user).key

        # 4. Create Supplier
        self.supplier1 = Taminotchi.objects.create(
            biznes=self.biznes1, nomi="Islomjon", telefon_raqam="+998991234567"
        )
        self.supplier2 = Taminotchi.objects.create(
            biznes=self.biznes2, nomi="Bekzod", telefon_raqam="+998997654321"
        )

        # 5. Create product
        self.p1 = Mahsulot.objects.create(
            biznes=self.biznes1, nomi="Armatura", olchov_birligi="dona",
            kelish_narxi=Decimal("40000.00"), sotish_narxi=Decimal("60000.00"), ulgurji_narx=Decimal("55000.00"), miqdori=10
        )
        MahsulotShtrixKod.objects.create(mahsulot=self.p1, kod="9948493123")
        DokonQoldiq.objects.create(mahsulot=self.p1, dokon=self.dokon1, miqdori=10)

    def test_supplier_order_workflow_manual(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.token1)

        # 1. Create draft order
        payload = {
            "taminotchi": self.supplier1.id,
            "dokon": self.dokon1.id,
            "qabul_qilish_sanasi": "2026-07-20",
            "elementlar": [
                {
                    "mahsulot": self.p1.id,
                    "miqdori": 100,
                    "kelish_narxi": "42000.00",
                    "ustama": "50.00",
                    "sotish_narxi": "63000.00",
                    "ulgurji_narx": "58000.00"
                }
            ]
        }
        response = self.client.post(reverse('supplier-order-list'), payload, format='json')
        if response.status_code != 201:
            print("MANUAL WORKFLOW CREATE ERROR:", response.content)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        order_id = response.data['id']
        self.assertEqual(response.data['holat'], 'qoralama')
        self.assertEqual(Decimal(response.data['umumiy_summa']), Decimal('4200000.00')) # 100 * 42000

        # 2. Confirm order (Rasmiylashtirish)
        confirm_url = reverse('supplier-order-confirm', kwargs={'pk': order_id})
        response = self.client.post(confirm_url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['holat'], 'rasmiylashtirilgan')

        # 3. Pay partial amount using cash (Naqd)
        pay_url = reverse('supplier-order-pay', kwargs={'pk': order_id})
        pay_payload = {
            "amount": "2000000.00",
            "tolov_turi": "naqd"
        }
        response = self.client.post(pay_url, pay_payload, format='json')
        if response.status_code != 200:
            print("PAYMENT WORKFLOW ERROR:", response.content)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        # 4. Check price differences (Order: 63000, Store: 60000)
        diff_url = reverse('supplier-order-price-differences', kwargs={'pk': order_id})
        response = self.client.get(diff_url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 1)
        self.assertEqual(Decimal(response.data[0]['buyurtma_narxi']), Decimal('63000.00'))

        # 5. Receive order with price overwrite = True
        receive_url = reverse('supplier-order-receive', kwargs={'pk': order_id})
        receive_payload = {
            "apply_new_prices": True
        }
        response = self.client.post(receive_url, receive_payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['holat'], 'qabul_qilingan')

        # 6. Verify inventory and prices
        self.p1.refresh_from_db()
        # Stock: 10 original + 100 received = 110
        self.assertEqual(self.p1.miqdori, 110)
        # Price overwritten to 63000
        self.assertEqual(self.p1.sotish_narxi, Decimal('63000.00'))
        self.assertEqual(self.p1.kelish_narxi, Decimal('42000.00'))

    def test_supplier_order_workflow_excel(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.token1)

        # 1. Create in-memory Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Shablon"
        headers = ["Nomi", "Shtrix-kod", "Buyurtmaga", "Kelish narxi", "Ustama %", "Sotuv narxi", "Ulgurji narx"]
        ws.append(headers)
        # We order a new product "Semun"
        ws.append(["Semun", "40783213123", "50", "10000", "50", "15000", "13000"])
        
        file_data = BytesIO()
        wb.save(file_data)
        file_data.seek(0)
        file_data.name = "buyurtma.xlsx"

        # 2. Create order via Excel upload
        payload = {
            "taminotchi": self.supplier1.id,
            "dokon": self.dokon1.id,
            "qabul_qilish_sanasi": "2026-07-20",
            "fayl": file_data
        }
        response = self.client.post(reverse('supplier-order-list'), payload, format='multipart')
        if response.status_code != 201:
            print("EXCEL WORKFLOW CREATE ERROR:", response.content)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        order_id = response.data['id']
        self.assertEqual(Decimal(response.data['umumiy_summa']), Decimal('500000.00')) # 50 * 10000

        # 3. Check new product "Semun" was auto-created in database
        semun = Mahsulot.objects.get(nomi="Semun")
        self.assertEqual(semun.sotish_narxi, Decimal('15000.00'))

    def test_multi_tenant_isolation(self):
        # Admin B queries supplier list
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.token2)
        response = self.client.get(reverse('taminotchi-list'))
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        # Should only see self.supplier2 (1 supplier) belonging to biznes2
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]['id'], self.supplier2.id)

        # Admin A queries supplier list
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.token1)
        response = self.client.get(reverse('taminotchi-list'))
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        # Should only see self.supplier1 (1 supplier) belonging to biznes1
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]['id'], self.supplier1.id)

    def test_excel_upload_product_limit_enforced(self):
        self.tarif.mahsulot_limiti = 1
        self.tarif.save()

        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.token1)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Shablon"
        headers = ["Nomi", "Shtrix-kod", "Buyurtmaga", "Kelish narxi", "Ustama %", "Sotuv narxi", "Ulgurji narx"]
        ws.append(headers)
        ws.append(["Semun Limit Test", "40783213999", "10", "10000", "50", "15000", "13000"])
        
        file_data = BytesIO()
        wb.save(file_data)
        file_data.seek(0)
        file_data.name = "buyurtma_limit.xlsx"

        payload = {
            "taminotchi": self.supplier1.id,
            "dokon": self.dokon1.id,
            "qabul_qilish_sanasi": "2026-07-20",
            "fayl": file_data
        }
        
        response = self.client.post(reverse('supplier-order-list'), payload, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_supplier_balance_and_returns(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.token1)

        # 1. Start with initial balance of 0 on supplier1
        self.assertEqual(self.supplier1.balans, Decimal('0.00'))

        # 2. Receive an order return to increase supplier balance
        payload_order = {
            "taminotchi": self.supplier1.id,
            "dokon": self.dokon1.id,
            "qabul_qilish_sanasi": "2026-07-20",
            "elementlar": [
                {
                    "mahsulot": self.p1.id,
                    "miqdori": 10,
                    "kelish_narxi": "40000.00",
                    "ustama": "50.00",
                    "sotish_narxi": "60000.00",
                    "ulgurji_narx": "55000.00"
                }
            ]
        }
        res_o = self.client.post(reverse('supplier-order-list'), payload_order, format='json')
        self.assertEqual(res_o.status_code, status.HTTP_201_CREATED)
        order_id = res_o.data['id']

        # Confirm and receive order
        self.client.post(reverse('supplier-order-confirm', kwargs={'pk': order_id}))
        self.client.post(reverse('supplier-order-receive', kwargs={'pk': order_id}), {"apply_new_prices": False}, format='json')

        # Now return 5 items of product self.p1 to supplier1
        payload_ret = {
            "order": order_id,
            "dokon": self.dokon1.id,
            "taminotchi": self.supplier1.id,
            "elementlar": [
                {
                    "mahsulot": self.p1.id,
                    "miqdori": 5,
                    "kelish_narxi": "40000.00"
                }
            ]
        }
        res_ret = self.client.post(reverse('supplier-order-return-list'), payload_ret, format='json')
        self.assertEqual(res_ret.status_code, status.HTTP_201_CREATED)
        ret_id = res_ret.data['id']

        # Confirm the return
        res_confirm = self.client.post(reverse('supplier-order-return-confirm', kwargs={'pk': ret_id}))
        self.assertEqual(res_confirm.status_code, status.HTTP_200_OK)

        # Check supplier1 balance increased by qaytarish_summasi (5 * 40000 = 200000)
        self.supplier1.refresh_from_db()
        self.assertEqual(self.supplier1.balans, Decimal('200000.00'))

        # 3. Create a second order, confirm it, and try paying using 'balans_postavshika'
        res_o2 = self.client.post(reverse('supplier-order-list'), payload_order, format='json')
        self.assertEqual(res_o2.status_code, status.HTTP_201_CREATED)
        order2_id = res_o2.data['id']
        self.client.post(reverse('supplier-order-confirm', kwargs={'pk': order2_id}))

        # Try to pay 300000.00 using balans_postavshika (we only have 200000.00)
        res_pay = self.client.post(reverse('supplier-order-pay', kwargs={'pk': order2_id}), {"amount": "300000.00", "tolov_turi": "balans_postavshika"}, format='json')
        self.assertEqual(res_pay.status_code, status.HTTP_400_BAD_REQUEST)

        # Pay 150000.00 using balans_postavshika (should succeed)
        res_pay2 = self.client.post(reverse('supplier-order-pay', kwargs={'pk': order2_id}), {"amount": "150000.00", "tolov_turi": "balans_postavshika"}, format='json')
        self.assertEqual(res_pay2.status_code, status.HTTP_200_OK)

        # Check supplier1 balance decreased (200000 - 150000 = 50000)
        self.supplier1.refresh_from_db()
        self.assertEqual(self.supplier1.balans, Decimal('50000.00'))

        # 4. Cancel the second order. The paid amount via balans_postavshika (150000) should be refunded to supplier balance
        res_cancel = self.client.post(reverse('supplier-order-cancel', kwargs={'pk': order2_id}))
        self.assertEqual(res_cancel.status_code, status.HTTP_200_OK)

        # Check supplier balance is refunded (50000 + 150000 = 200000)
        self.supplier1.refresh_from_db()
        self.assertEqual(self.supplier1.balans, Decimal('200000.00'))

    def test_supplier_order_payment_status_filtering(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.token1)
        
        payload = {
            "taminotchi": self.supplier1.id,
            "dokon": self.dokon1.id,
            "qabul_qilish_sanasi": "2026-07-20",
            "elementlar": [
                {
                    "mahsulot": self.p1.id,
                    "miqdori": 10,
                    "kelish_narxi": "40000.00",
                    "ustama": "50.00",
                    "sotish_narxi": "60000.00",
                    "ulgurji_narx": "55000.00"
                }
            ]
        }
        
        # Order 1: Unpaid (tolanmagan)
        r1 = self.client.post(reverse('supplier-order-list'), payload, format='json')
        self.assertEqual(r1.status_code, status.HTTP_201_CREATED)
        self.assertEqual(r1.data['tolov_status'], 'tolanmagan')
        
        # Order 2: Partially paid (qisman_tolangan)
        r2 = self.client.post(reverse('supplier-order-list'), payload, format='json')
        self.assertEqual(r2.status_code, status.HTTP_201_CREATED)
        o2_id = r2.data['id']
        self.client.post(reverse('supplier-order-confirm', kwargs={'pk': o2_id}))
        self.client.post(reverse('supplier-order-pay', kwargs={'pk': o2_id}), {"amount": "100000.00", "tolov_turi": "naqd"}, format='json')
        
        # Order 3: Fully paid (tolangan)
        r3 = self.client.post(reverse('supplier-order-list'), payload, format='json')
        self.assertEqual(r3.status_code, status.HTTP_201_CREATED)
        o3_id = r3.data['id']
        self.client.post(reverse('supplier-order-confirm', kwargs={'pk': o3_id}))
        self.client.post(reverse('supplier-order-pay', kwargs={'pk': o3_id}), {"amount": "400000.00", "tolov_turi": "naqd"}, format='json')

        # List all
        res = self.client.get(reverse('supplier-order-list'))
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        
        # List unpaid
        res = self.client.get(reverse('supplier-order-list'), {"to_lov_status": "tolanmagan"})
        self.assertEqual(len(res.data), 1)
        self.assertEqual(res.data[0]['id'], r1.data['id'])

        # List partially paid
        res = self.client.get(reverse('supplier-order-list'), {"to_lov_status": "qisman_tolangan"})
        self.assertEqual(len(res.data), 1)
        self.assertEqual(res.data[0]['id'], o2_id)
        self.assertIsNotNone(res.data[0]['tolangan_vaqt'])

        # List fully paid
        res = self.client.get(reverse('supplier-order-list'), {"to_lov_status": "tolangan"})
        self.assertEqual(len(res.data), 1)
        self.assertEqual(res.data[0]['id'], o3_id)
        self.assertIsNotNone(res.data[0]['tolangan_vaqt'])

    def test_supplier_order_excel_export(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.token1)
        response = self.client.get(reverse('supplier-order-list'), {"export": "excel"})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    def test_supplier_order_template(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.token1)
        response = self.client.get(reverse('supplier-order-template'))
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        wb = openpyxl.load_workbook(filename=BytesIO(response.content), data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        self.assertIn("Nomi", headers)
        self.assertIn("Shtrix-kod", headers)
        self.assertIn("Buyurtmaga", headers)
        self.assertIn("Kelish narxi", headers)

    def test_supplier_order_search_and_tolov_status(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.token1)
        
        payload = {
            "nomi": "Search Order Test",
            "taminotchi": self.supplier1.id,
            "dokon": self.dokon1.id,
            "qabul_qilish_sanasi": "2026-07-20",
            "elementlar": [
                {
                    "mahsulot": self.p1.id,
                    "miqdori": 10,
                    "kelish_narxi": "40000.00",
                    "sotish_narxi": "60000.00"
                }
            ]
        }
        res = self.client.post(reverse('supplier-order-list'), payload, format='json')
        self.assertEqual(res.status_code, status.HTTP_201_CREATED)
        o_id = res.data['id']
        
        response = self.client.get(reverse('supplier-order-list'), {"search": str(o_id)})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        ids = [item['id'] for item in response.data]
        self.assertIn(o_id, ids)
        
        response = self.client.get(reverse('supplier-order-list'), {"tolov_status": "tolanmagan"})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        ids = [item['id'] for item in response.data]
        self.assertIn(o_id, ids)

