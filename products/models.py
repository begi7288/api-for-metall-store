import random
import csv
import openpyxl
import os
from io import BytesIO
from django.db import models
from django.core.validators import MinValueValidator, MaxValueValidator, RegexValidator
from django.core.exceptions import ValidationError
from decimal import Decimal

def validate_image_size(value):
    if not value:
        return
    # LOW-5: Fayl kengaytmasini tekshirish
    import os
    ext = os.path.splitext(value.name)[1].lower()
    allowed_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp']
    if ext not in allowed_extensions:
        raise ValidationError(f"Faqat rasm fayllari ruxsat etiladi ({', '.join(allowed_extensions)}).")
    limit = 5 * 1024 * 1024
    if value.size > limit:
        raise ValidationError("Rasm hajmi 5MB dan oshmasligi kerak.")

def validate_import_file_extension(value):
    if not value:
        return
    ext = os.path.splitext(value.name)[1].lower()
    valid_extensions = ['.xlsx', '.xls', '.csv']
    if ext not in valid_extensions:
        raise ValidationError("Faqat Excel (.xlsx, .xls) yoki CSV (.csv) fayllari ruxsat etiladi.")

def validate_import_file_size(value):
    if not value:
        return
    limit = 10 * 1024 * 1024
    if value.size > limit:
        raise ValidationError("Import fayli hajmi 10MB dan oshmasligi kerak.")


from temirdokon_v1.models import BaseModel
from user.models import Biznes

class Characteristic(BaseModel):
    name = models.CharField(max_length=255)
    value = models.CharField(max_length=255)

    def __str__(self):
        return f"{self.name}: {self.value}"

class Mahsulot(BaseModel):
    OLCHOV_CHOICES = (
        ('kg', 'Kilogramm'),
        ('dona', 'Dona'),
        ('metr', 'Metr'),
        ('litr', 'Litr'),
    )
    biznes = models.ForeignKey(Biznes, on_delete=models.CASCADE, related_name='mahsulotlar', null=True, blank=True)
    nomi = models.CharField(max_length=255)
    olchov_birligi = models.CharField(max_length=50, choices=OLCHOV_CHOICES)
    kelish_narxi = models.DecimalField(max_digits=12, decimal_places=2)
    ustama = models.DecimalField(max_digits=5, decimal_places=2, default=0.00, validators=[MinValueValidator(0.00), MaxValueValidator(100.00)])
    sotish_narxi = models.DecimalField(max_digits=12, decimal_places=2, blank=True, null=True)
    miqdori = models.PositiveIntegerField(default=0)
    ogohlantirish = models.PositiveIntegerField(default=0)
    is_active = models.BooleanField(default=True)
    toifa = models.CharField(max_length=100, blank=True, null=True, default="Mavjud emas")
    brend = models.CharField(max_length=100, blank=True, null=True, default="Mavjud emas")
    taminotchi = models.ForeignKey('Taminotchi', on_delete=models.SET_NULL, null=True, blank=True, related_name='mahsulotlar')
    erkin_narx = models.BooleanField(default=False)
    ulgurji_narx = models.DecimalField(max_digits=12, decimal_places=2, default=Decimal('0.00'), blank=True, null=True)
    tavsif = models.TextField(blank=True, null=True)
    characteristics = models.ManyToManyField(Characteristic, blank=True, related_name='mahsulotlar')
    dokonlar = models.ManyToManyField('Dokon', through='DokonQoldiq', related_name='mahsulotlar')

    def clean(self):
        super().clean()

        # 1. Boundary validations
        if self.kelish_narxi is not None and self.kelish_narxi <= 0:
            raise ValidationError({'kelish_narxi': "Kelish narxi 0 dan katta bo'lishi kerak."})

        if self.miqdori is not None and self.miqdori < 0:
            raise ValidationError({'miqdori': "Miqdor manfiy bo'lishi mumkin emas."})

        if self.ogohlantirish is not None and self.ogohlantirish < 0:
            raise ValidationError({'ogohlantirish': "Ogohlantirish miqdori manfiy bo'lishi mumkin emas."})

        # 3. Price & Markup Logic
        if self.kelish_narxi is not None:
            if self.sotish_narxi is not None:
                if self.sotish_narxi < self.kelish_narxi:
                    raise ValidationError({'sotish_narxi': "Sotish narxi kelish narxidan kichik bo'lishi mumkin emas."})
                self.ustama = (((self.sotish_narxi - self.kelish_narxi) / self.kelish_narxi) * Decimal('100.00')).quantize(Decimal('0.01'))
                if self.ustama < 0 or self.ustama > 100:
                    raise ValidationError({'sotish_narxi': "Sotish narxi ustamasi 0% va 100% oralig'ida bo'lishi kerak."})
            else:
                if self.ustama is None:
                    self.ustama = Decimal('0.00')
                self.sotish_narxi = (self.kelish_narxi * (Decimal('1.00') + self.ustama / Decimal('100.00'))).quantize(Decimal('0.01'))

    def save(self, *args, **kwargs):
        is_new = not self.pk
        super().save(*args, **kwargs)
        if is_new:
            custom_barcodes = getattr(self, '_custom_barcodes', None)
            if custom_barcodes:
                for code in custom_barcodes:
                    MahsulotShtrixKod.objects.create(mahsulot=self, kod=code)
            else:
                code = self.generate_unique_barcode()
                MahsulotShtrixKod.objects.create(mahsulot=self, kod=code)

    def generate_unique_barcode(self):
        while True:
            code = "".join([str(random.randint(0, 9)) for _ in range(13)])
            if not MahsulotShtrixKod.objects.filter(kod=code).exists():
                return code

    def __str__(self):
        return f"{self.nomi} ({self.miqdori} {self.get_olchov_birligi_display()})"

    @property
    def shtrix_kod(self):
        first_kod = self.shtrix_kodlar.first()
        return first_kod.kod if first_kod else None

    @shtrix_kod.setter
    def shtrix_kod(self, value):
        if value:
            if self.pk:
                first_kod = self.shtrix_kodlar.first()
                if first_kod:
                    first_kod.kod = value
                    first_kod.save()
                else:
                    MahsulotShtrixKod.objects.create(mahsulot=self, kod=value)
            else:
                self._custom_barcodes = [value]

    @property
    def kam_qoldi(self):
        return any(q.miqdori <= q.ogohlantirish for q in self.qoldiqlar.all())


class MahsulotRasm(BaseModel):
    mahsulot = models.ForeignKey(Mahsulot, on_delete=models.CASCADE, related_name='rasmlar')
    rasm = models.ImageField(upload_to='mahsulotlar/', validators=[validate_image_size])

    def __str__(self):
        return f"Rasm: {self.mahsulot.nomi}"


class MahsulotShtrixKod(BaseModel):
    mahsulot = models.ForeignKey(Mahsulot, on_delete=models.CASCADE, related_name='shtrix_kodlar')
    kod = models.CharField(max_length=50, unique=True, validators=[
        RegexValidator(r'^\d+$', "Shtrix kod faqat raqamlardan iborat bo'lishi shart.")
    ])

    def __str__(self):
        return f"{self.mahsulot.nomi}: {self.kod}"


class Import(BaseModel):
    HOLAT_CHOICES = (
        ('kutilmoqda', 'Kutilmoqda'),
        ('yakunlangan', 'Yakunlangan'),
        ('xatolik', 'Xatolik'),
    )
    TURI_CHOICES = (
        ('kirim', 'Kirim'),
        ('qoldiq_kirimi', 'Qoldiq kirimi'),
    )
    biznes = models.ForeignKey(Biznes, on_delete=models.CASCADE, related_name='importlar', null=True, blank=True)
    dokon = models.ForeignKey('Dokon', on_delete=models.CASCADE, related_name='importlar', null=True, blank=True)
    nomi = models.CharField(max_length=255)
    fayl = models.FileField(
        upload_to='importlar/',
        validators=[validate_import_file_extension, validate_import_file_size]
    )
    holat = models.CharField(max_length=20, choices=HOLAT_CHOICES, default='kutilmoqda')
    import_turi = models.CharField(max_length=20, choices=TURI_CHOICES, default='kirim')
    shtrixkod_generatsiya_qilish = models.BooleanField(default=False)
    miqdori = models.PositiveIntegerField(default=0)
    kelish_summasi = models.DecimalField(max_digits=15, decimal_places=2, default=0.00)
    sotish_summasi = models.DecimalField(max_digits=15, decimal_places=2, default=0.00)
    sotuvlar_taraqqiyoti = models.FloatField(default=0.0)
    elementlar = models.JSONField(default=list, blank=True)
    yaratgan_xodim = models.ForeignKey('user.Xodim', on_delete=models.SET_NULL, null=True, related_name='yaratgan_importlar', blank=True)
    yakunlagan_xodim = models.ForeignKey('user.Xodim', on_delete=models.SET_NULL, null=True, related_name='yakunlagan_importlar', blank=True)

    def clean(self):
        super().clean()
        if not self.nomi:
            raise ValidationError({'nomi': "Import nomi kiritilishi shart."})

    def save(self, *args, **kwargs):
        is_new = not self.pk
        super().save(*args, **kwargs)
        if is_new and self.fayl:
            try:
                self.parse_and_save_elements()
            except ValidationError:
                # Re-raise django ValidationErrors directly so they are returned as API bad requests
                self.holat = 'xatolik'
                super().save(update_fields=['holat'])
                raise
            except Exception as e:
                self.holat = 'xatolik'
                super().save(update_fields=['holat'])
                raise ValidationError(f"Faylni tahlil qilishda xatolik yuz berdi: {str(e)}")

    def parse_and_save_elements(self):
        file_name = self.fayl.name
        self.fayl.seek(0)
        content = self.fayl.read()

        rows = []
        if file_name.endswith('.xlsx') or file_name.endswith('.xls'):
            wb = openpyxl.load_workbook(filename=BytesIO(content), data_only=True, read_only=True)
            sheet = wb.active
            for row in sheet.iter_rows(values_only=True):
                if any(x is not None for x in row):
                    rows.append([str(x) if x is not None else "" for x in row])
        else:
            # Assume CSV
            try:
                decoded = content.decode('utf-8')
            except UnicodeDecodeError:
                decoded = content.decode('latin-1')
            reader = csv.reader(decoded.splitlines())
            for row in reader:
                if any(x != "" for x in row):
                    rows.append(row)

        if not rows:
            raise ValidationError("Yuklangan fayl bo'sh yoki uni o'qib bo'lmadi.")

        headers = [str(h).lower().strip() for h in rows[0]]
        col_mapping = {}
        for idx, h in enumerate(headers):
            if any(k in h for k in ['nomi', 'name', 'наименование', 'tovar']):
                col_mapping['nomi'] = idx
            elif any(k in h for k in ['shtrix', 'barcode', 'barkod', 'баркод', 'kod', 'код']):
                col_mapping['shtrix_kod'] = idx
            elif any(k in h for k in ['miqdor', 'qty', 'kol', 'кол']):
                col_mapping['miqdori'] = idx
            elif any(k in h for k in ['kelish', 'cost', 'поставки']):
                col_mapping['kelish_narxi'] = idx
            elif any(k in h for k in ['sotish', 'retail', 'продажи', 'sotuv', 'розничная']):
                col_mapping['sotish_narxi'] = idx
            elif any(k in h for k in ['birlik', 'unit', 'единица', 'o\'lchov']):
                col_mapping['olchov_birligi'] = idx
            elif any(k in h for k in ['toifa', 'category', 'категория']):
                col_mapping['toifa'] = idx
            elif any(k in h for k in ['brend', 'brand', 'бренд']):
                col_mapping['brend'] = idx
            elif any(k in h for k in ['taminotchi', 'supplier', 'поставщик', 'yetkazib']):
                col_mapping['taminotchi'] = idx
            elif any(k in h for k in ['tavsif', 'description', 'описание']):
                col_mapping['tavsif'] = idx
            elif any(k in h for k in ['xususiyat', 'characteristic', 'feature']):
                col_mapping['characteristics'] = idx

        # Find active custom fields mapping dynamically
        from products.models import XususiyatMaydoni
        active_fields = XususiyatMaydoni.objects.filter(biznes=self.biznes, is_active=True)
        active_field_names = [f.nomi.lower().strip() for f in active_fields]
        
        custom_col_mapping = {}
        for idx, h in enumerate(headers):
            if h in active_field_names:
                matching_field = next(f for f in active_fields if f.nomi.lower().strip() == h)
                custom_col_mapping[matching_field.nomi] = idx

        # Fallback to column indices if auto-detect fails
        if 'nomi' not in col_mapping and len(headers) > 0:
            col_mapping['nomi'] = 0
        if 'shtrix_kod' not in col_mapping and len(headers) > 1:
            col_mapping['shtrix_kod'] = 1
        if 'miqdori' not in col_mapping and len(headers) > 2:
            col_mapping['miqdori'] = 2
        if 'kelish_narxi' not in col_mapping and len(headers) > 3:
            col_mapping['kelish_narxi'] = 3
        if 'sotish_narxi' not in col_mapping and len(headers) > 4:
            col_mapping['sotish_narxi'] = 4
        if 'olchov_birligi' not in col_mapping and len(headers) > 5:
            col_mapping['olchov_birligi'] = 5

        parsed_items = []
        total_qty = 0
        total_kelish = Decimal('0.00')
        total_sotish = Decimal('0.00')
        row_errors = []

        for idx, row in enumerate(rows[1:], start=2):
            if not any(str(x).strip() for x in row if x is not None):
                continue

            def get_val(key, default=""):
                idx = col_mapping.get(key)
                if idx is not None and idx < len(row):
                    val = row[idx]
                    return str(val).strip() if val is not None else default
                return default

            nomi = get_val('nomi')
            if not nomi:
                row_errors.append(f"Qator {idx}: Mahsulot nomi kiritilishi shart.")
                continue

            shtrix_kod = get_val('shtrix_kod')
            
            try:
                raw_qty = get_val('miqdori', '0')
                qty_str = "".join(filter(str.isdigit, raw_qty))
                miqdori = int(qty_str) if qty_str else 0
                if '-' in raw_qty:
                    miqdori = -miqdori
            except ValueError:
                miqdori = 0

            try:
                cost_str = "".join(c for c in get_val('kelish_narxi', '0') if c.isdigit() or c == '.')
                kelish_narxi = float(cost_str) if cost_str else 0.0
            except ValueError:
                kelish_narxi = 0.0

            try:
                sell_str = "".join(c for c in get_val('sotish_narxi', '0') if c.isdigit() or c == '.')
                sotish_narxi = float(sell_str) if sell_str else 0.0
            except ValueError:
                sotish_narxi = 0.0

            # Determine if product exists
            product_exists = False
            if shtrix_kod:
                product_exists = Mahsulot.objects.filter(biznes=self.biznes, shtrix_kodlar__kod=shtrix_kod).exists()
            if not product_exists:
                product_exists = Mahsulot.objects.filter(biznes=self.biznes, nomi=nomi).exists()

            # Content validations
            if miqdori < 0:
                row_errors.append(f"Qator {idx}: Miqdor manfiy bo'lishi mumkin emas ({miqdori}).")

            if not product_exists:
                if kelish_narxi <= 0:
                    row_errors.append(f"Qator {idx}: Yangi mahsulot '{nomi}' uchun kelish narxi kiritilishi shart.")
            else:
                if kelish_narxi < 0:
                    row_errors.append(f"Qator {idx}: Kelish narxi manfiy bo'lishi mumkin emas.")

            if sotish_narxi > 0:
                compare_cost = kelish_narxi
                if compare_cost <= 0 and product_exists:
                    existing_p = None
                    if shtrix_kod:
                        existing_p = Mahsulot.objects.filter(biznes=self.biznes, shtrix_kodlar__kod=shtrix_kod).first()
                    if not existing_p:
                        existing_p = Mahsulot.objects.filter(biznes=self.biznes, nomi=nomi).first()
                    if existing_p:
                        compare_cost = float(existing_p.kelish_narxi)
                
                if compare_cost > 0 and sotish_narxi < compare_cost:
                    row_errors.append(f"Qator {idx}: Sotish narxi kelish narxidan ({compare_cost}) kichik bo'lishi mumkin emas.")

            unit_str = get_val('olchov_birligi').lower()
            olchov_birligi = 'dona'
            if any(u in unit_str for u in ['кг', 'kg', 'kilogramm']):
                olchov_birligi = 'kg'
            elif any(u in unit_str for u in ['м', 'metr', 'm']):
                olchov_birligi = 'metr'
            elif any(u in unit_str for u in ['л', 'litr', 'l']):
                olchov_birligi = 'litr'

            total_qty += miqdori
            total_kelish += Decimal(str(kelish_narxi)) * miqdori
            total_sotish += Decimal(str(sotish_narxi)) * miqdori

            toifa = get_val('toifa')
            brend = get_val('brend')
            taminotchi_nomi = get_val('taminotchi')
            tavsif = get_val('tavsif')

            characteristics_list = []
            
            # Read dynamic custom field values
            for field_name, col_idx in custom_col_mapping.items():
                if col_idx < len(row):
                    val = row[col_idx]
                    val_str = str(val).strip() if val is not None else ""
                    if val_str:
                        characteristics_list.append({
                            'name': field_name,
                            'value': val_str
                        })

            # Read fallback comma-separated characteristics if present
            chars_str = get_val('characteristics', '')
            if chars_str:
                parts = chars_str.split(',')
                for part in parts:
                    if ':' in part:
                        c_name, c_val = part.split(':', 1)
                        c_name = c_name.strip()
                        c_val = c_val.strip()
                        if c_name and c_val:
                            characteristics_list.append({
                                'name': c_name,
                                'value': c_val
                            })

            parsed_items.append({
                'nomi': nomi,
                'shtrix_kod': shtrix_kod,
                'miqdori': miqdori,
                'kelish_narxi': kelish_narxi,
                'sotish_narxi': sotish_narxi,
                'olchov_birligi': olchov_birligi,
                'toifa': toifa,
                'brend': brend,
                'taminotchi_nomi': taminotchi_nomi,
                'tavsif': tavsif,
                'characteristics': characteristics_list
            })

        if row_errors:
            raise ValidationError({'fayl': row_errors})

        self.miqdori = total_qty
        self.kelish_summasi = total_kelish
        self.sotish_summasi = total_sotish
        self.elementlar = parsed_items
        if self.pk:
            super().save(update_fields=['miqdori', 'kelish_summasi', 'sotish_summasi', 'elementlar'])

    def confirm_and_execute(self, executor_xodim=None):
        if self.holat != 'kutilmoqda':
            raise ValidationError("Ushbu import allaqachon yakunlangan yoki bekor qilingan.")

        for item in self.elementlar:
            shtrix_kod = item.get('shtrix_kod')
            nomi = item.get('nomi')
            miqdori = item.get('miqdori', 0)
            kelish_narxi = Decimal(str(item.get('kelish_narxi', 0.0)))
            sotish_narxi = Decimal(str(item.get('sotish_narxi', 0.0)))
            olchov_birligi = item.get('olchov_birligi', 'dona')

            toifa = item.get('toifa')
            brend = item.get('brend')
            taminotchi_nomi = item.get('taminotchi_nomi')
            tavsif = item.get('tavsif')

            taminotchi_obj = None
            if taminotchi_nomi:
                taminotchi_obj, created = Taminotchi.objects.get_or_create(
                    biznes=self.biznes,
                    nomi=taminotchi_nomi
                )

            product = None
            if shtrix_kod:
                try:
                    product = Mahsulot.objects.get(biznes=self.biznes, shtrix_kodlar__kod=shtrix_kod)
                except Mahsulot.DoesNotExist:
                    pass

            if not product:
                try:
                    product = Mahsulot.objects.get(biznes=self.biznes, nomi=nomi)
                except (Mahsulot.DoesNotExist, Mahsulot.MultipleObjectsReturned):
                    pass

            # Get characteristics list from item
            characteristics_list = item.get('characteristics', [])
            characteristics_objs = []
            for char_data in characteristics_list:
                c_name = char_data.get('name')
                c_val = char_data.get('value')
                if c_name and c_val:
                    char_obj, created = Characteristic.objects.get_or_create(name=c_name, value=c_val)
                    characteristics_objs.append(char_obj)

            if product:
                if self.dokon:
                    qoldiq, created = DokonQoldiq.objects.get_or_create(mahsulot=product, dokon=self.dokon)
                    if self.import_turi == 'kirim':
                        qoldiq.miqdori += miqdori
                    else:
                        qoldiq.miqdori = miqdori
                    qoldiq.save()
                    product.miqdori = sum(q.miqdori for q in product.qoldiqlar.all())
                else:
                    if self.import_turi == 'kirim':
                        product.miqdori += miqdori
                    else:
                        product.miqdori = miqdori

                if kelish_narxi > 0:
                    product.kelish_narxi = kelish_narxi
                if sotish_narxi > 0:
                    product.sotish_narxi = sotish_narxi

                if kelish_narxi > 0 and sotish_narxi > 0:
                    product.ustama = (((sotish_narxi - kelish_narxi) / kelish_narxi) * Decimal('100.00')).quantize(Decimal('0.01'))

                if toifa:
                    product.toifa = toifa
                if brend:
                    product.brend = brend
                if tavsif:
                    product.tavsif = tavsif
                if taminotchi_obj:
                    product.taminotchi = taminotchi_obj

                product.save()
                if characteristics_objs:
                    product.characteristics.add(*characteristics_objs)
                item['shtrix_kod'] = product.shtrix_kod
            else:
                code = shtrix_kod
                if not code and self.shtrixkod_generatsiya_qilish:
                    code = Mahsulot().generate_unique_barcode()

                ustama = Decimal('0.00')
                if kelish_narxi > 0 and sotish_narxi > 0:
                    ustama = (((sotish_narxi - kelish_narxi) / kelish_narxi) * Decimal('100.00')).quantize(Decimal('0.01'))

                if self.biznes and self.biznes.tarif:
                    limit = self.biznes.tarif.mahsulot_limiti
                    if Mahsulot.objects.filter(biznes=self.biznes).count() >= limit:
                        raise ValidationError(f"Tarif rejangiz bo'yicha mahsulotlar soni limiti ({limit}) tugagan. Yangi mahsulot yaratib bo'lmaydi.")

                new_prod = Mahsulot(
                    biznes=self.biznes,
                    nomi=nomi,
                    olchov_birligi=olchov_birligi,
                    kelish_narxi=kelish_narxi,
                    sotish_narxi=sotish_narxi or None,
                    ustama=ustama,
                    miqdori=0,
                    toifa=toifa or "Mavjud emas",
                    brend=brend or "Mavjud emas",
                    taminotchi=taminotchi_obj,
                    tavsif=tavsif
                )
                if code:
                    new_prod._custom_barcodes = [code]
                new_prod.save()
                
                if self.dokon:
                    DokonQoldiq.objects.create(mahsulot=new_prod, dokon=self.dokon, miqdori=miqdori)
                    new_prod.miqdori = miqdori
                    new_prod.save(update_fields=['miqdori'])
                else:
                    new_prod.miqdori = miqdori
                    new_prod.save(update_fields=['miqdori'])

                if characteristics_objs:
                    new_prod.characteristics.set(characteristics_objs)
                item['shtrix_kod'] = new_prod.shtrix_kod

        self.holat = 'yakunlangan'
        if executor_xodim:
            self.yakunlagan_xodim = executor_xodim
        self.save()

    def __str__(self):
        return f"Import: {self.nomi} ({self.get_holat_display()})"


class Dokon(BaseModel):
    biznes = models.ForeignKey(Biznes, on_delete=models.CASCADE, related_name='dokonlar', null=True, blank=True)
    nomi = models.CharField(max_length=255, unique=True)
    tavsif = models.TextField(blank=True, null=True)

    def __str__(self):
        return self.nomi


class DokonQoldiq(BaseModel):
    mahsulot = models.ForeignKey(Mahsulot, on_delete=models.CASCADE, related_name='qoldiqlar')
    dokon = models.ForeignKey(Dokon, on_delete=models.CASCADE, related_name='qoldiqlar')
    miqdori = models.PositiveIntegerField(default=0)
    ogohlantirish = models.PositiveIntegerField(default=0)

    class Meta:
        unique_together = ('mahsulot', 'dokon')

    def __str__(self):
        return f"{self.mahsulot.nomi} - {self.dokon.nomi}: {self.miqdori}"


class Transfer(BaseModel):
    HOLAT_CHOICES = (
        ('kutilmoqda', 'Kutilmoqda'),
        ('yakunlangan', 'Yakunlangan'),
        ('xatolik', 'Xatolik'),
    )
    biznes = models.ForeignKey(Biznes, on_delete=models.CASCADE, related_name='transferlar', null=True, blank=True)
    nomi = models.CharField(max_length=255)
    dokondan = models.ForeignKey(Dokon, on_delete=models.CASCADE, related_name='yuborilgan_transferlar')
    dokonga = models.ForeignKey(Dokon, on_delete=models.CASCADE, related_name='qabul_qilingan_transferlar')
    fayl = models.FileField(
        upload_to='transferlar/',
        validators=[validate_import_file_extension, validate_import_file_size],
        blank=True,
        null=True
    )
    holat = models.CharField(max_length=20, choices=HOLAT_CHOICES, default='kutilmoqda')
    miqdori = models.PositiveIntegerField(default=0)
    summa = models.DecimalField(max_digits=15, decimal_places=2, default=0.00)
    elementlar = models.JSONField(default=list, blank=True)
    yaratgan_xodim = models.ForeignKey('user.Xodim', on_delete=models.SET_NULL, null=True, related_name='yaratgan_transferlar', blank=True)
    qabul_qilgan_xodim = models.ForeignKey('user.Xodim', on_delete=models.SET_NULL, null=True, related_name='qabul_qilgan_transferlar', blank=True)

    def clean(self):
        super().clean()
        if not self.nomi:
            raise ValidationError({'nomi': "Transfer nomi kiritilishi shart."})
        
        if self.dokondan_id and self.dokonga_id and self.dokondan_id == self.dokonga_id:
            raise ValidationError("Jo'natuvchi va qabul qiluvchi do'konlar bir xil bo'lishi mumkin emas.")

        # If not from file and elementlar is provided, validate items
        if not self.fayl and self.elementlar:
            total_qty = 0
            total_sum = Decimal('0.00')
            validated_items = []
            errors = []
            for idx, item in enumerate(self.elementlar, start=1):
                nomi = item.get('nomi')
                shtrix_kod = item.get('shtrix_kod')
                miqdori = item.get('miqdori', 0)
                
                try:
                    miqdori = int(miqdori)
                except (ValueError, TypeError):
                    errors.append(f"Element {idx}: Miqdor butun son bo'lishi kerak.")
                    continue
                
                if miqdori <= 0:
                    errors.append(f"Element {idx}: O'tkazilayotgan miqdor 0 dan katta bo'lishi kerak.")
                    continue
                
                product = None
                if shtrix_kod:
                    product = Mahsulot.objects.filter(biznes=self.biznes, shtrix_kodlar__kod=shtrix_kod).first()
                if not product and nomi:
                    product = Mahsulot.objects.filter(biznes=self.biznes, nomi=nomi).first()
                
                if not product:
                    errors.append(f"Element {idx}: Mahsulot bazada mavjud emas (Nomi: '{nomi}', Shtrix kod: '{shtrix_kod}').")
                else:
                    qoldiq = DokonQoldiq.objects.filter(mahsulot=product, dokon=self.dokondan).first()
                    sender_qty = qoldiq.miqdori if qoldiq else 0
                    if sender_qty < miqdori:
                        errors.append(f"Element {idx}: Omborda yetarli miqdor mavjud emas. So'ralgan: {miqdori}, Mavjud (yuboruvchi do'kon): {sender_qty}.")
                    validated_items.append({
                        'nomi': product.nomi,
                        'shtrix_kod': product.shtrix_kod,
                        'miqdori': miqdori,
                        'olchov_birligi': product.olchov_birligi
                    })
                    total_qty += miqdori
                    total_sum += Decimal(str(product.sotish_narxi or product.kelish_narxi or 0.0)) * miqdori
            
            if errors:
                raise ValidationError({'elementlar': errors})
            
            self.miqdori = total_qty
            self.summa = total_sum
            self.elementlar = validated_items

    def save(self, *args, **kwargs):
        is_new = not self.pk
        super().save(*args, **kwargs)
        if is_new and self.fayl:
            try:
                self.parse_and_save_elements()
            except ValidationError:
                self.holat = 'xatolik'
                super().save(update_fields=['holat'])
                raise
            except Exception as e:
                self.holat = 'xatolik'
                super().save(update_fields=['holat'])
                raise ValidationError(f"Faylni tahlil qilishda xatolik yuz berdi: {str(e)}")

    def parse_and_save_elements(self):
        if not self.fayl:
            return
        
        file_name = self.fayl.name
        self.fayl.seek(0)
        content = self.fayl.read()

        rows = []
        if file_name.endswith('.xlsx') or file_name.endswith('.xls'):
            wb = openpyxl.load_workbook(filename=BytesIO(content), data_only=True, read_only=True)
            sheet = wb.active
            for row in sheet.iter_rows(values_only=True):
                if any(x is not None for x in row):
                    rows.append([str(x) if x is not None else "" for x in row])
        else:
            # Assume CSV
            try:
                decoded = content.decode('utf-8')
            except UnicodeDecodeError:
                decoded = content.decode('latin-1')
            reader = csv.reader(decoded.splitlines())
            for row in reader:
                if any(x != "" for x in row):
                    rows.append(row)

        if not rows:
            raise ValidationError("Yuklangan fayl bo'sh yoki uni o'qib bo'lmadi.")

        headers = [str(h).lower().strip() for h in rows[0]]
        col_mapping = {}
        for idx, h in enumerate(headers):
            if any(k in h for k in ['nomi', 'name', 'наименование', 'tovar']):
                col_mapping['nomi'] = idx
            elif any(k in h for k in ['shtrix', 'barcode', 'барkod', 'баркод', 'kod', 'код']):
                col_mapping['shtrix_kod'] = idx
            elif any(k in h for k in ['miqdor', 'qty', 'kol', 'кол']):
                col_mapping['miqdori'] = idx
            elif any(k in h for k in ['birlik', 'unit', 'единица', 'o\'lchov']):
                col_mapping['olchov_birligi'] = idx

        # Fallback to column indices if auto-detect fails
        if 'nomi' not in col_mapping and len(headers) > 0:
            col_mapping['nomi'] = 0
        if 'shtrix_kod' not in col_mapping and len(headers) > 1:
            col_mapping['shtrix_kod'] = 1
        if 'miqdori' not in col_mapping and len(headers) > 2:
            col_mapping['miqdori'] = 2
        if 'olchov_birligi' not in col_mapping and len(headers) > 3:
            col_mapping['olchov_birligi'] = 3

        parsed_items = []
        total_qty = 0
        total_sum = Decimal('0.00')
        row_errors = []

        for idx, row in enumerate(rows[1:], start=2):
            if not any(str(x).strip() for x in row if x is not None):
                continue

            def get_val(key, default=""):
                idx = col_mapping.get(key)
                if idx is not None and idx < len(row):
                    val = row[idx]
                    return str(val).strip() if val is not None else default
                return default

            nomi = get_val('nomi')
            shtrix_kod = get_val('shtrix_kod')
            
            if not nomi and not shtrix_kod:
                row_errors.append(f"Qator {idx}: Mahsulot nomi yoki shtrix kodi kiritilishi shart.")
                continue

            try:
                raw_qty = get_val('miqdori', '0')
                qty_str = "".join(filter(str.isdigit, raw_qty))
                miqdori = int(qty_str) if qty_str else 0
                if '-' in raw_qty:
                    miqdori = -miqdori
            except ValueError:
                miqdori = 0

            if miqdori <= 0:
                row_errors.append(f"Qator {idx}: O'tkazilayotgan miqdor 0 dan katta bo'lishi kerak ({miqdori}).")

            product = None
            if shtrix_kod:
                product = Mahsulot.objects.filter(biznes=self.biznes, shtrix_kodlar__kod=shtrix_kod).first()
            if not product and nomi:
                product = Mahsulot.objects.filter(biznes=self.biznes, nomi=nomi).first()

            if not product:
                row_errors.append(f"Qator {idx}: Mahsulot bazada mavjud emas (Nomi: '{nomi}', Shtrix kod: '{shtrix_kod}').")
            else:
                qoldiq = DokonQoldiq.objects.filter(mahsulot=product, dokon=self.dokondan).first()
                sender_qty = qoldiq.miqdori if qoldiq else 0
                if sender_qty < miqdori:
                    row_errors.append(f"Qator {idx}: Omborda yetarli miqdor mavjud emas. So'ralgan: {miqdori}, Mavjud (yuboruvchi do'kon): {sender_qty}.")
                total_sum += Decimal(str(product.sotish_narxi or product.kelish_narxi or 0.0)) * miqdori

            unit_str = get_val('olchov_birligi').lower()
            olchov_birligi = 'dona'
            if any(u in unit_str for u in ['кг', 'kg', 'kilogramm']):
                olchov_birligi = 'kg'
            elif any(u in unit_str for u in ['м', 'metr', 'm']):
                olchov_birligi = 'metr'
            elif any(u in unit_str for u in ['л', 'litr', 'l']):
                olchov_birligi = 'litr'

            total_qty += miqdori
            parsed_items.append({
                'nomi': product.nomi if product else nomi,
                'shtrix_kod': product.shtrix_kod if product else shtrix_kod,
                'miqdori': miqdori,
                'olchov_birligi': product.olchov_birligi if product else olchov_birligi
            })

        if row_errors:
            raise ValidationError({'fayl': row_errors})

        self.miqdori = total_qty
        self.summa = total_sum
        self.elementlar = parsed_items
        if self.pk:
            super().save(update_fields=['miqdori', 'summa', 'elementlar'])

    def confirm_and_execute(self, executor_xodim=None):
        if self.holat != 'kutilmoqda':
            raise ValidationError("Ushbu transfer allaqachon yakunlangan yoki bekor qilingan.")

        # Deduct inventory
        for item in self.elementlar:
            shtrix_kod = item.get('shtrix_kod')
            nomi = item.get('nomi')
            miqdori = item.get('miqdori', 0)

            product = None
            if shtrix_kod:
                product = Mahsulot.objects.filter(biznes=self.biznes, shtrix_kodlar__kod=shtrix_kod).first()
            if not product and nomi:
                product = Mahsulot.objects.filter(biznes=self.biznes, nomi=nomi).first()

            if not product:
                raise ValidationError(f"Mahsulot bazada topilmadi: {nomi or shtrix_kod}")
            
            sender_qoldiq = DokonQoldiq.objects.filter(mahsulot=product, dokon=self.dokondan).first()
            sender_qty = sender_qoldiq.miqdori if sender_qoldiq else 0
            if sender_qty < miqdori:
                raise ValidationError(f"Omborda yetarli miqdor mavjud emas (Mahsulot: {product.nomi}, Mavjud: {sender_qty}, So'ralgan: {miqdori}).")
            
            sender_qoldiq.miqdori -= miqdori
            sender_qoldiq.save()
            
            receiver_qoldiq, created = DokonQoldiq.objects.get_or_create(mahsulot=product, dokon=self.dokonga)
            receiver_qoldiq.miqdori += miqdori
            receiver_qoldiq.save()
            
            product.miqdori = sum(q.miqdori for q in product.qoldiqlar.all())
            product.save(update_fields=['miqdori'])

        self.holat = 'yakunlangan'
        if executor_xodim:
            self.qabul_qilgan_xodim = executor_xodim
        self.save()

    def __str__(self):
        return f"Transfer: {self.nomi} ({self.get_holat_display()})"

class Taminotchi(BaseModel):
    biznes = models.ForeignKey(Biznes, on_delete=models.CASCADE, related_name='taminotchilar', null=True, blank=True)
    nomi = models.CharField(max_length=255)
    telefon_raqam = models.CharField(max_length=50, blank=True, null=True)
    telefonlar = models.JSONField(default=list, blank=True)
    standart_ustama = models.DecimalField(max_digits=5, decimal_places=2, default=Decimal('0.00'))
    eslatma = models.TextField(blank=True, null=True)

    # Requisites (Rekvizitlar)
    yuridik_nomi = models.CharField(max_length=255, blank=True, null=True)
    yuridik_manzil = models.CharField(max_length=500, blank=True, null=True)
    mamlakat = models.CharField(max_length=100, blank=True, null=True)
    pochta_indeksi = models.CharField(max_length=50, blank=True, null=True)
    bank_hisob_raqami = models.CharField(max_length=100, blank=True, null=True)
    bank_nomi_filiali = models.CharField(max_length=255, blank=True, null=True)
    inn = models.CharField(max_length=50, blank=True, null=True)
    mfo = models.CharField(max_length=50, blank=True, null=True)

    balans = models.DecimalField(max_digits=15, decimal_places=2, default=Decimal('0.00'))

    def __str__(self):
        return self.nomi

class WriteOff(BaseModel):
    SABABI_CHOICES = (
        ('defekt', "Yaroqsiz (Defekt)"),
        ('yoqotish', "Yo'qotish (Kamomad)"),
        ('katalogdan_ochirish', "Katalogdan o'chirish"),
        ('saralash_xatosi', "Saralash xatoligini to'g'rilash"),
        ('inventarizatsiya', "Inventarizatsiya"),
        ('boshqa', "Boshqa"),
    )
    HOLAT_CHOICES = (
        ('qoralama', 'Qoralama'),
        ('yakunlangan', 'Yakunlangan'),
        ('bekor_qilingan', 'Bekor qilingan'),
    )
    biznes = models.ForeignKey(Biznes, on_delete=models.CASCADE, related_name='hisobdan_chiqarishlar', null=True, blank=True)
    dokon = models.ForeignKey(Dokon, on_delete=models.CASCADE, related_name='hisobdan_chiqarishlar')
    nomi = models.CharField(max_length=255)
    sababi = models.CharField(max_length=50, choices=SABABI_CHOICES, default='boshqa')
    holat = models.CharField(max_length=20, choices=HOLAT_CHOICES, default='qoralama')
    miqdori = models.PositiveIntegerField(default=0)
    kelish_summasi = models.DecimalField(max_digits=15, decimal_places=2, default=Decimal('0.00'))
    sotish_summasi = models.DecimalField(max_digits=15, decimal_places=2, default=Decimal('0.00'))
    fayl = models.FileField(upload_to='write_offs/', null=True, blank=True)
    fayldan_hisobdan_chiqarish = models.BooleanField(default=False)
    yaratgan_xodim = models.ForeignKey('user.Xodim', on_delete=models.SET_NULL, null=True, related_name='yaratgan_hisobdan_chiqarishlar', blank=True)
    tasdiqlagan_xodim = models.ForeignKey('user.Xodim', on_delete=models.SET_NULL, null=True, related_name='tasdiqlagan_hisobdan_chiqarishlar', blank=True)

    def clean(self):
        super().clean()
        if not self.nomi:
            raise ValidationError({'nomi': "Hisobdan chiqarish nomi kiritilishi shart."})
        if self.biznes and self.dokon and self.dokon.biznes != self.biznes:
            raise ValidationError("Do'kon sizning biznesingizga tegishli emas.")

    def confirm_and_execute(self, executor_xodim=None):
        if self.holat != 'qoralama':
            raise ValidationError("Faqat qoralama holatidagi hujjatni tasdiqlash mumkin.")
            
        if not self.elementlar.exists():
            raise ValidationError("Hisobdan chiqarish uchun kamida bitta mahsulot kiritilgan bo'lishi shart.")

        for item in self.elementlar.all():
            product = item.mahsulot
            qty = item.miqdori

            # Find dokon qoldiq
            try:
                qoldiq = DokonQoldiq.objects.get(mahsulot=product, dokon=self.dokon)
            except DokonQoldiq.DoesNotExist:
                raise ValidationError(f"'{product.nomi}' mahsuloti uchun '{self.dokon.nomi}' do'konida qoldiq topilmadi.")

            if qoldiq.miqdori < qty:
                raise ValidationError(f"'{product.nomi}' uchun do'konda yetarli qoldiq mavjud emas. Do'kondagi qoldiq: {qoldiq.miqdori}, So'ralgan: {qty}.")

            # Deduct stock
            qoldiq.miqdori -= qty
            qoldiq.save()

            # Recalculate total product count
            product.miqdori = sum(q.miqdori for q in product.qoldiqlar.all())
            product.save(update_fields=['miqdori'])

        self.holat = 'yakunlangan'
        if executor_xodim:
            self.tasdiqlagan_xodim = executor_xodim
        self.save()

    def bekor_qilish(self):
        if self.holat != 'qoralama':
            raise ValidationError("Faqat qoralama holatidagi hujjatni bekor qilish mumkin.")
        self.holat = 'bekor_qilingan'
        self.save()

    def __str__(self):
        return f"{self.nomi} ({self.get_holat_display()})"


class WriteOffItem(BaseModel):
    write_off = models.ForeignKey(WriteOff, on_delete=models.CASCADE, related_name='elementlar')
    mahsulot = models.ForeignKey(Mahsulot, on_delete=models.PROTECT, related_name='hisobdan_chiqarish_elementlari')
    miqdori = models.PositiveIntegerField()
    kelish_narxi = models.DecimalField(max_digits=12, decimal_places=2, default=Decimal('0.00'))
    sotish_narxi = models.DecimalField(max_digits=12, decimal_places=2, default=Decimal('0.00'))

    def clean(self):
        super().clean()
        if self.miqdori is not None and self.miqdori <= 0:
            raise ValidationError({'miqdori': "Hisobdan chiqariladigan miqdor 0 dan katta bo'lishi shart."})
        if self.write_off and self.write_off.biznes and self.mahsulot.biznes != self.write_off.biznes:
            raise ValidationError("Tanlangan mahsulot sizning biznesingizga tegishli emas.")
        if self.write_off and self.write_off.dokon:
            qoldiq_exists = DokonQoldiq.objects.filter(mahsulot=self.mahsulot, dokon=self.write_off.dokon).exists()
            if not qoldiq_exists:
                raise ValidationError(f"'{self.mahsulot.nomi}' mahsuloti ushbu do'konda mavjud emas (qoldiq jadvalida topilmadi).")

    def save(self, *args, **kwargs):
        self.clean()
        if self.mahsulot:
            if not self.kelish_narxi or self.kelish_narxi == Decimal('0.00'):
                self.kelish_narxi = self.mahsulot.kelish_narxi or Decimal('0.00')
            if not self.sotish_narxi or self.sotish_narxi == Decimal('0.00'):
                self.sotish_narxi = self.mahsulot.sotish_narxi or Decimal('0.00')
        super().save(*args, **kwargs)
        self.recalculate_totals()

    def delete(self, *args, **kwargs):
        write_off = self.write_off
        super().delete(*args, **kwargs)
        # Trigger parent recalculation
        write_off.miqdori = write_off.elementlar.aggregate(total=models.Sum('miqdori'))['total'] or 0
        write_off.kelish_summasi = write_off.elementlar.aggregate(total=models.Sum(models.F('miqdori') * models.F('kelish_narxi')))['total'] or Decimal('0.00')
        write_off.sotish_summasi = write_off.elementlar.aggregate(total=models.Sum(models.F('miqdori') * models.F('sotish_narxi')))['total'] or Decimal('0.00')
        write_off.save()

    def recalculate_totals(self):
        write_off = self.write_off
        write_off.miqdori = write_off.elementlar.aggregate(total=models.Sum('miqdori'))['total'] or 0
        write_off.kelish_summasi = write_off.elementlar.aggregate(total=models.Sum(models.F('miqdori') * models.F('kelish_narxi')))['total'] or Decimal('0.00')
        write_off.sotish_summasi = write_off.elementlar.aggregate(total=models.Sum(models.F('miqdori') * models.F('sotish_narxi')))['total'] or Decimal('0.00')
        write_off.save()

    def __str__(self):
        return f"{self.mahsulot.nomi} - {self.miqdori}"


class XususiyatMaydoni(BaseModel):
    biznes = models.ForeignKey(Biznes, on_delete=models.CASCADE, related_name='xususiyat_maydonlari', null=True, blank=True)
    nomi = models.CharField(max_length=255)
    tur = models.CharField(max_length=50, default='matn')
    is_active = models.BooleanField(default=True)

    def __str__(self):
        return f"{self.nomi} (Matn)"


class Toplam(BaseModel):
    biznes = models.ForeignKey(Biznes, on_delete=models.CASCADE, related_name='toplamlar', null=True, blank=True)
    nomi = models.CharField(max_length=255, blank=True, null=True)
    dokon = models.ForeignKey('Dokon', on_delete=models.CASCADE, related_name='toplamlar')
    holat = models.CharField(max_length=50, default='qoralama', choices=(('qoralama', 'Qoralama'), ('yakunlangan', 'Tasdiqlangan')))
    miqdori = models.PositiveIntegerField(default=0)
    summa = models.DecimalField(max_digits=15, decimal_places=2, default=Decimal('0.00'))
    yaratgan_xodim = models.ForeignKey('user.Xodim', on_delete=models.SET_NULL, null=True, blank=True, related_name='yaratgan_toplamlar')

    def __str__(self):
        return f"To'plam {self.id} - {self.dokon.nomi}"

    def confirm_and_execute(self, executor_xodim=None):
        if self.holat != 'qoralama':
            raise ValidationError("Faqat qoralama holatidagi to'plamni tasdiqlash mumkin.")
            
        if not self.elementlar.exists():
            raise ValidationError("To'plamga kamida bitta mahsulot kiritilgan bo'lishi shart.")

        for item in self.elementlar.all():
            product = item.mahsulot
            qty = item.miqdori

            # Find or create dokon qoldiq
            qoldiq, created = DokonQoldiq.objects.get_or_create(
                mahsulot=product, 
                dokon=self.dokon,
                defaults={'miqdori': 0, 'ogohlantirish': 0}
            )
            # Add stock
            qoldiq.miqdori += qty
            qoldiq.save()

            # Recalculate total product count
            product.miqdori = sum(q.miqdori for q in product.qoldiqlar.all())
            product.save(update_fields=['miqdori'])

        self.holat = 'yakunlangan'
        self.save()


class ToplamElement(BaseModel):
    toplam = models.ForeignKey(Toplam, on_delete=models.CASCADE, related_name='elementlar')
    mahsulot = models.ForeignKey('Mahsulot', on_delete=models.CASCADE, related_name='toplam_elementlari')
    miqdori = models.PositiveIntegerField(default=0)
    kelish_narxi = models.DecimalField(max_digits=12, decimal_places=2, default=Decimal('0.00'))
    sotish_narxi = models.DecimalField(max_digits=12, decimal_places=2, default=Decimal('0.00'))

    def clean(self):
        super().clean()
        if self.miqdori is not None and self.miqdori <= 0:
            raise ValidationError({'miqdori': "Miqdor 0 dan katta bo'lishi shart."})

    def save(self, *args, **kwargs):
        self.clean()
        if self.mahsulot:
            if not self.kelish_narxi or self.kelish_narxi == Decimal('0.00'):
                self.kelish_narxi = self.mahsulot.kelish_narxi or Decimal('0.00')
            if not self.sotish_narxi or self.sotish_narxi == Decimal('0.00'):
                self.sotish_narxi = self.mahsulot.sotish_narxi or Decimal('0.00')
        super().save(*args, **kwargs)
        self.recalculate_totals()

    def delete(self, *args, **kwargs):
        toplam = self.toplam
        super().delete(*args, **kwargs)
        toplam.miqdori = toplam.elementlar.aggregate(total=models.Sum('miqdori'))['total'] or 0
        toplam.summa = toplam.elementlar.aggregate(total=models.Sum(models.F('miqdori') * models.F('kelish_narxi')))['total'] or Decimal('0.00')
        toplam.save()

    def recalculate_totals(self):
        toplam = self.toplam
        toplam.miqdori = toplam.elementlar.aggregate(total=models.Sum('miqdori'))['total'] or 0
        toplam.summa = toplam.elementlar.aggregate(total=models.Sum(models.F('miqdori') * models.F('kelish_narxi')))['total'] or Decimal('0.00')
        toplam.save()


class YorliqShablon(BaseModel):
    biznes = models.ForeignKey(Biznes, on_delete=models.CASCADE, related_name='yorliq_shablonlari', null=True, blank=True)
    nomi = models.CharField(max_length=255)
    eni = models.FloatField(default=40.0)      # Width in mm
    uzunlik = models.FloatField(default=20.0)  # Length/Height in mm
    shtrixkod_formati = models.CharField(max_length=50, default='CODE128')
    xususiyatlar = models.JSONField(default=list, blank=True) # List of fields to display, e.g. ["nomi", "sotish_narxi", "shtrix_kod", "Qalinligi"]
    is_active = models.BooleanField(default=True)

    def __str__(self):
        return f"Shablon: {self.nomi} ({self.eni}x{self.uzunlik} mm)"