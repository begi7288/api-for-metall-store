from rest_framework.test import APITestCase
from rest_framework import status
from django.urls import reverse
from rest_framework.authtoken.models import Token
from decimal import Decimal
from django.core.files.uploadedfile import SimpleUploadedFile
import openpyxl
from io import BytesIO

from user.models import Biznes, Xodim, Tarif
from products.models import Dokon, Mahsulot, XususiyatMaydoni, DokonQoldiq

class ImportAPITestCase(APITestCase):
    def setUp(self):
        self.tarif = Tarif.objects.create(nomi="Pro", dokon_limiti=5, mahsulot_limiti=100, xodim_limiti=5)
        self.biznes = Biznes.objects.create(nomi="Import Biznes", egasi_ism="Importer", tarif=self.tarif)
        
        from django.contrib.auth.models import User
        self.u1 = User.objects.create_user(username="importer_u", password="password123")
        self.x1 = Xodim.objects.create(
            user=self.u1, ism="Ali", familiya="Valiyev", telefon_raqam="+998909999999", 
            parol="secret123", jinsi="erkak", biznes=self.biznes, rol="admin"
        )
        self.t1 = Token.objects.create(user=self.u1).key
        self.dokon = Dokon.objects.create(biznes=self.biznes, nomi="Main Store")

        self.field1 = XususiyatMaydoni.objects.create(biznes=self.biznes, nomi="Qalinligi", tur="matn", is_active=True)

        self.list_url = reverse('import-list')
        self.template_url = reverse('import-template')

        # Create cashier (restricted role)
        self.u_cashier = User.objects.create_user(username="importer_c", password="password123")
        self.cashier_xodim = Xodim.objects.create(
            user=self.u_cashier,
            biznes=self.biznes,
            ism="Cashier", familiya="Restricted", telefon_raqam="+998908888888",
            parol="cashierpassword123", rol="sotuvchi", jinsi="ayol"
        )
        self.cashier_token = Token.objects.create(user=self.u_cashier).key

    def test_import_csv_flow(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.t1)
        csv_data = (
            "Nomi,Shtrix-kod,Miqdori,Kelish narxi,Sotish narxi,O'lchov birligi,Xususiyatlar\n"
            "Armatura,9948493123123,300,40000.00,60000.00,Штука,\"Brend: Quvasoy, Qalinligi: 1.5mm\"\n"
        )
        mock_file = SimpleUploadedFile("armatura.csv", csv_data.encode('utf-8'), content_type="text/csv")
        
        payload = {
            "nomi": "Armatura kirim",
            "fayl": mock_file,
            "import_turi": "kirim",
            "dokon": self.dokon.id
        }
        
        response = self.client.post(self.import_list_url, payload, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data['holat'], 'yakunlangan')
        self.assertEqual(response.data['miqdori'], 300)
        self.assertEqual(float(response.data['kelish_summasi']), 12000000.0)
        
        product = Mahsulot.objects.get(shtrix_kodlar__kod="9948493123123")
        self.assertEqual(product.nomi, "Armatura")
        self.assertEqual(product.characteristics.count(), 2)
        self.assertTrue(product.characteristics.filter(name="Brend", value="Quvasoy").exists())
        self.assertTrue(product.characteristics.filter(name="Qalinligi", value="1.5mm").exists())

    def test_import_invalid_file_extension(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.t1)
        mock_file = SimpleUploadedFile("invalid.txt", b"some plain text content", content_type="text/plain")
        
        payload = {
            "nomi": "Invalid Extension Import",
            "fayl": mock_file
        }
        response = self.client.post(self.import_list_url, payload, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_import_invalid_row_values(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.t1)
        csv_data = (
            "Nomi,Shtrix-kod,Miqdori,Kelish narxi,Sotish narxi,O'lchov birligi\n"
            "Armatura,9948493123123,300,50000.00,40000.00,Штука\n"
            "Gips,9948493124124,-10,10000.00,12000.00,кг\n"
        )
        mock_file = SimpleUploadedFile("bad_values.csv", csv_data.encode('utf-8'), content_type="text/csv")
        
        payload = {
            "nomi": "Bad Values Import",
            "fayl": mock_file
        }
        response = self.client.post(self.import_list_url, payload, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('fayl', response.data['errors'])
        self.assertTrue(any("Qator 2" in err or "Qator 3" in err for err in response.data['errors']['fayl']))

    def test_import_cashier_access_denied(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.cashier_token)
        csv_data = "Nomi,Shtrix-kod,Miqdori,Kelish narxi,Sotish narxi,O'lchov birligi\nArmatura,9948493123123,300,40000.00,60000.00,Штука\n"
        mock_file = SimpleUploadedFile("armatura.csv", csv_data.encode('utf-8'), content_type="text/csv")
        
        payload = {
            "nomi": "Cashier Try Import",
            "fayl": mock_file
        }
        response = self.client.post(self.import_list_url, payload, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_download_template_contains_dynamic_fields(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.t1)
        response = self.client.get(self.template_url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        wb = openpyxl.load_workbook(filename=BytesIO(response.content), data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        self.assertIn("Nomi", headers)
        self.assertIn("Shtrix-kod", headers)
        self.assertIn("Qalinligi", headers)
        self.assertNotIn("Artikul", headers)

    def test_import_execution_success(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.t1)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([
            "Nomi", "Shtrix-kod", "Miqdori", "Kelish narxi", "Sotish narxi",
            "Toifa", "Brend", "O'lchov birligi", "Yetkazib beruvchi", "Tavsif", "Qalinligi"
        ])
        ws.append([
            "Imported Tovar", "9876543210123", "30", "5000.00", "7500.00",
            "Metallar", "Premium", "dona", "Premium supplier", "Great metal", "5mm"
        ])
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        fayl = SimpleUploadedFile("test_import.xlsx", excel_file.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        
        payload = {
            "nomi": "Test Excel Import",
            "dokon": self.dokon.id,
            "fayl": fayl,
            "import_turi": "kirim",
            "shtrixkod_generatsiya_qilish": False
        }
        
        response = self.client.post(self.import_list_url, payload, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data['holat'], 'yakunlangan')
        
        product = Mahsulot.objects.get(biznes=self.biznes, nomi="Imported Tovar")
        self.assertEqual(product.kelish_narxi, Decimal("5000.00"))
        self.assertEqual(product.sotish_narxi, Decimal("7500.00"))
        self.assertEqual(product.toifa, "Metallar")
        self.assertEqual(product.brend, "Premium")
        self.assertEqual(product.tavsif, "Great metal")
        
        self.assertEqual(product.taminotchi.nomi, "Premium supplier")
        self.assertEqual(product.taminotchi.biznes, self.biznes)
        
        dq = DokonQoldiq.objects.get(mahsulot=product, dokon=self.dokon)
        self.assertEqual(dq.miqdori, 30)
        self.assertEqual(product.miqdori, 30)
        
        char = product.characteristics.filter(name="Qalinligi", value="5mm").first()
        self.assertIsNotNone(char)

    @property
    def import_list_url(self):
        return self.list_url
