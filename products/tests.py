from django.test import TestCase
from django.core.exceptions import ValidationError
from decimal import Decimal
from rest_framework.test import APITestCase
from rest_framework import status
from django.urls import reverse
from rest_framework.authtoken.models import Token
from user.models import Xodim, Biznes
from .models import Mahsulot, Import, Characteristic, MahsulotRasm, MahsulotShtrixKod, Dokon, DokonQoldiq


class MahsulotModelTest(TestCase):
    def test_product_characteristics(self):
        # Create Characteristic instances
        char1 = Characteristic.objects.create(name="Rang", value="Qizil")
        char2 = Characteristic.objects.create(name="Material", value="Temir")
        
        # Create Mahsulot
        product = Mahsulot.objects.create(
            nomi="Truba",
            olchov_birligi="metr",
            kelish_narxi=Decimal("15000.00"),
            ustama=Decimal("10.00"),
            miqdori=20
        )
        
        # Associate characteristics
        product.characteristics.add(char1, char2)
        
        # Assert relationships
        self.assertEqual(product.characteristics.count(), 2)
        self.assertIn(char1, product.characteristics.all())
        self.assertIn(char2, product.characteristics.all())

    def test_auto_calculate_sotish_narxi(self):
        # Providing kelish_narxi and ustama, sotish_narxi should be calculated automatically
        product = Mahsulot(
            nomi="Gips",
            olchov_birligi="kg",
            kelish_narxi=Decimal("100.00"),
            ustama=Decimal("20.00"),
            miqdori=5,
            ogohlantirish=2
        )
        product.save()
        self.assertEqual(product.sotish_narxi, Decimal("120.00"))
        self.assertIsNotNone(product.shtrix_kod)
        self.assertEqual(len(product.shtrix_kod), 13)

    def test_auto_calculate_ustama(self):
        # Providing kelish_narxi and sotish_narxi, ustama should be calculated automatically
        product = Mahsulot(
            nomi="Tsement",
            olchov_birligi="kg",
            kelish_narxi=Decimal("100.00"),
            sotish_narxi=Decimal("125.00"),
            miqdori=5
        )
        product.save()
        self.assertEqual(product.ustama, Decimal("25.00"))

    def test_selling_at_a_loss_validation(self):
        # Creating product with selling price less than cost price should raise ValidationError
        product = Mahsulot(
            nomi="Lopata",
            olchov_birligi="dona",
            kelish_narxi=Decimal("100.00"),
            sotish_narxi=Decimal("90.00")
        )
        with self.assertRaises(ValidationError):
            product.full_clean()

    def test_invalid_barcode_validation(self):
        # Non-numeric barcode should raise ValidationError
        product = Mahsulot.objects.create(
            nomi="Vedro",
            olchov_birligi="dona",
            kelish_narxi=Decimal("50.00"),
            ustama=Decimal("10.00"),
            miqdori=5
        )
        invalid_bc = MahsulotShtrixKod(mahsulot=product, kod="abc123xyz")
        with self.assertRaises(ValidationError):
            invalid_bc.full_clean()

    def test_negative_values_validation(self):
        # Negative cost price should raise ValidationError
        product = Mahsulot(
            nomi="Shlang",
            olchov_birligi="metr",
            kelish_narxi=Decimal("-10.00"),
            ustama=Decimal("10.00")
        )
        with self.assertRaises(ValidationError):
            product.full_clean()

    def test_image_size_validation(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        # 6MB mock file
        large_file = SimpleUploadedFile("large_image.jpg", b"x" * (6 * 1024 * 1024))
        product = Mahsulot.objects.create(
            nomi="Gips",
            olchov_birligi="kg",
            kelish_narxi=Decimal("100.00"),
            ustama=Decimal("20.00"),
            miqdori=5
        )
        prod_image = MahsulotRasm(mahsulot=product, rasm=large_file)
        with self.assertRaises(ValidationError):
            prod_image.full_clean()


class MahsulotAPITestCase(APITestCase):
    def setUp(self):
        self.list_url = reverse('mahsulot-list')
        
        # Create Biznes and Dokon
        self.biznes = Biznes.objects.create(nomi="Test Biznes", egasi_ism="Owner")
        self.dokon = Dokon.objects.create(biznes=self.biznes, nomi="Test Dokon")

        # Create an admin/omborchi employee to bypass read-only permissions
        self.manager_xodim = Xodim.objects.create(
            biznes=self.biznes,
            ism="Manager", familiya="Testov", telefon_raqam="+998907777777",
            parol="managerpassword123", rol="omborchi", jinsi="erkak"
        )
        self.manager_token = Token.objects.create(user=self.manager_xodim.user).key
        
        # Create a sotuvchi employee (restricted)
        self.cashier_xodim = Xodim.objects.create(
            biznes=self.biznes,
            ism="Cashier", familiya="Testova", telefon_raqam="+998906666666",
            parol="cashierpassword123", rol="sotuvchi", jinsi="ayol"
        )
        self.cashier_token = Token.objects.create(user=self.cashier_xodim.user).key

        # Pre-create some products for searching/filtering tests
        self.p1 = Mahsulot.objects.create(
            biznes=self.biznes,
            nomi="Shifer Plastik", olchov_birligi="dona",
            kelish_narxi=Decimal("50000.00"), ustama=Decimal("20.00"),
            miqdori=100, ogohlantirish=10
        )
        DokonQoldiq.objects.create(mahsulot=self.p1, dokon=self.dokon, miqdori=100, ogohlantirish=10)

        self.p2 = Mahsulot.objects.create(
            biznes=self.biznes,
            nomi="Massa Gips", olchov_birligi="kg",
            kelish_narxi=Decimal("10000.00"), ustama=Decimal("10.00"),
            miqdori=5, ogohlantirish=10  # Low stock
        )
        DokonQoldiq.objects.create(mahsulot=self.p2, dokon=self.dokon, miqdori=5, ogohlantirish=10)

    def test_create_product_api_success(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.manager_token)
        payload = {
            "nomi": "Kran",
            "olchov_birligi": "dona",
            "kelish_narxi": "20000.00",
            "ustama": "15.00",
            "qoldiqlar": [
                {
                    "dokon": self.dokon.id,
                    "miqdori": 10,
                    "ogohlantirish": 2
                }
            ]
        }
        response = self.client.post(self.list_url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(Decimal(response.data['sotish_narxi']), Decimal("23000.00"))

    def test_create_product_api_unauthorized_for_cashier(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.cashier_token)
        payload = {
            "nomi": "Kran",
            "olchov_birligi": "dona",
            "kelish_narxi": "20000.00",
            "ustama": "15.00",
            "miqdori": 10
        }
        response = self.client.post(self.list_url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_read_products_allowed_for_cashier(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.cashier_token)
        response = self.client.get(self.list_url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 2)

    def test_retrieve_product_detail(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.cashier_token)
        detail_url = reverse('mahsulot-detail', kwargs={'pk': self.p1.pk})
        response = self.client.get(detail_url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['nomi'], "Shifer Plastik")

    def test_update_product_success(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.manager_token)
        detail_url = reverse('mahsulot-detail', kwargs={'pk': self.p1.pk})
        payload = {"nomi": "Shifer Plastik Yangi", "kelish_narxi": "50000.00", "sotish_narxi": "65000.00"}
        response = self.client.patch(detail_url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['nomi'], "Shifer Plastik Yangi")

    def test_delete_product_success(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.manager_token)
        detail_url = reverse('mahsulot-detail', kwargs={'pk': self.p1.pk})
        response = self.client.delete(detail_url)
        self.assertEqual(response.status_code, status.HTTP_204_NO_CONTENT)
        self.assertEqual(Mahsulot.objects.count(), 1)

    def test_filter_by_olchov_birligi(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.manager_token)
        response = self.client.get(self.list_url, {'olchov_birligi': 'kg'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]['nomi'], "Massa Gips")

    def test_search_by_name(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.manager_token)
        response = self.client.get(self.list_url, {'search': 'Plastik'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]['nomi'], "Shifer Plastik")

    def test_custom_filter_kam_qoldi(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.manager_token)
        # Verify that kam_qoldi=true returns only Massa Gips (miqdori=5 <= ogohlantirish=10)
        response = self.client.get(self.list_url, {'kam_qoldi': 'true'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]['nomi'], "Massa Gips")

    def test_product_api_with_characteristics(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.manager_token)
        char = Characteristic.objects.create(name="Kafolat", value="1 yil")
        self.p1.characteristics.add(char)
        
        detail_url = reverse('mahsulot-detail', kwargs={'pk': self.p1.pk})
        response = self.client.get(detail_url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn('characteristics', response.data)
        self.assertEqual(len(response.data['characteristics']), 1)
        self.assertEqual(response.data['characteristics'][0]['name'], "Kafolat")
        self.assertEqual(response.data['characteristics'][0]['value'], "1 yil")

    def test_create_product_with_characteristics(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.manager_token)
        char1 = Characteristic.objects.create(name="Rang", value="Qora")
        char2 = Characteristic.objects.create(name="Kafolat", value="1 yil")
        payload = {
            "nomi": "Dazmol",
            "olchov_birligi": "dona",
            "kelish_narxi": "150000.00",
            "ustama": "20.00",
            "characteristics": [char1.id, char2.id],
            "qoldiqlar": [
                {
                    "dokon": self.dokon.id,
                    "miqdori": 10,
                    "ogohlantirish": 2
                }
            ]
        }
        response = self.client.post(self.list_url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(len(response.data['characteristics']), 2)
        self.assertEqual(response.data['characteristics'][0]['name'], "Rang")
        self.assertEqual(response.data['characteristics'][0]['value'], "Qora")
        self.assertEqual(response.data['characteristics'][1]['name'], "Kafolat")
        self.assertEqual(response.data['characteristics'][1]['value'], "1 yil")

    def test_create_product_invalid_characteristics(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.manager_token)
        payload = {
            "nomi": "Dazmol",
            "olchov_birligi": "dona",
            "kelish_narxi": "150000.00",
            "ustama": "20.00",
            "miqdori": 10,
            "characteristics": [99999]  # Non-existent ID
        }
        response = self.client.post(self.list_url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_grouped_characteristics_api(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.manager_token)
        Characteristic.objects.create(name="Brend", value="Quvasoy")
        Characteristic.objects.create(name="Brend", value="Samsung")
        Characteristic.objects.create(name="Yetkazib beruvchi", value="Husniddin")
        
        grouped_url = reverse('characteristic-grouped')
        response = self.client.get(grouped_url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn('Brend', response.data)
        self.assertIn('Yetkazib beruvchi', response.data)
        self.assertEqual(len(response.data['Brend']), 2)
        self.assertEqual(len(response.data['Yetkazib beruvchi']), 1)

    def test_create_product_with_multiple_barcodes(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.manager_token)
        payload = {
            "nomi": "Dazmol",
            "olchov_birligi": "dona",
            "kelish_narxi": "150000.00",
            "ustama": "20.00",
            "shtrix_kod": ["12345678", "9876543210123"],
            "qoldiqlar": [
                {
                    "dokon": self.dokon.id,
                    "miqdori": 10,
                    "ogohlantirish": 2
                }
            ]
        }
        response = self.client.post(self.list_url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(len(response.data['shtrix_kod']), 2)
        self.assertIn("12345678", response.data['shtrix_kod'])
        self.assertIn("9876543210123", response.data['shtrix_kod'])

    def test_create_product_with_comma_separated_barcodes(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.manager_token)
        payload = {
            "nomi": "Mikroto'lqinli Pech",
            "olchov_birligi": "dona",
            "kelish_narxi": "500000.00",
            "ustama": "15.00",
            "shtrix_kod": "11112222, 33334444, 55556666",
            "qoldiqlar": [
                {
                    "dokon": self.dokon.id,
                    "miqdori": 5,
                    "ogohlantirish": 1
                }
            ]
        }
        response = self.client.post(self.list_url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(len(response.data['shtrix_kod']), 3)
        self.assertIn("11112222", response.data['shtrix_kod'])
        self.assertIn("33334444", response.data['shtrix_kod'])
        self.assertIn("55556666", response.data['shtrix_kod'])

    def test_create_product_with_duplicate_barcodes(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.manager_token)
        # Create a product with a barcode first
        payload1 = {
            "nomi": "Dazmol 1",
            "olchov_birligi": "dona",
            "kelish_narxi": "150000.00",
            "ustama": "20.00",
            "shtrix_kod": ["12345678"],
            "qoldiqlar": [
                {
                    "dokon": self.dokon.id,
                    "miqdori": 10,
                    "ogohlantirish": 2
                }
            ]
        }
        self.client.post(self.list_url, payload1, format='json')

        # Creating another product with the same barcode should fail
        payload2 = {
            "nomi": "Dazmol 2",
            "olchov_birligi": "dona",
            "kelish_narxi": "150000.00",
            "ustama": "20.00",
            "shtrix_kod": ["12345678"],
            "qoldiqlar": [
                {
                    "dokon": self.dokon.id,
                    "miqdori": 10,
                    "ogohlantirish": 2
                }
            ]
        }
        response = self.client.post(self.list_url, payload2, format='json')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('shtrix_kod', response.data['errors'])

    def test_search_by_multiple_barcodes(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.manager_token)
        # Add multiple barcodes to p1
        self.p1.shtrix_kodlar.all().delete()
        MahsulotShtrixKod.objects.create(mahsulot=self.p1, kod="1111111111111")
        MahsulotShtrixKod.objects.create(mahsulot=self.p1, kod="2222222222222")

        # Search by second barcode
        response = self.client.get(self.list_url, {'search': '2222222222222'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]['nomi'], "Shifer Plastik")

    def test_create_product_with_multiple_images(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.manager_token)
        from django.core.files.uploadedfile import SimpleUploadedFile
        from PIL import Image
        from io import BytesIO

        img = Image.new('RGB', (1, 1), color='red')
        buf = BytesIO()
        img.save(buf, format='JPEG')
        jpeg_data = buf.getvalue()

        img1 = SimpleUploadedFile("img1.jpg", jpeg_data, content_type="image/jpeg")
        img2 = SimpleUploadedFile("img2.jpg", jpeg_data, content_type="image/jpeg")
        img3 = SimpleUploadedFile("img3.jpg", jpeg_data, content_type="image/jpeg")
        
        payload = {
            "nomi": "Dazmol",
            "olchov_birligi": "dona",
            "kelish_narxi": "150000.00",
            "ustama": "20.00",
            "rasm": [img1, img2, img3],
            "qoldiqlar": '[{"dokon": %d, "miqdori": 10, "ogohlantirish": 2}]' % self.dokon.id
        }
        response = self.client.post(self.list_url, payload, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(len(response.data['rasm']), 3)

    def test_create_product_with_too_many_images(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.manager_token)
        from django.core.files.uploadedfile import SimpleUploadedFile
        from PIL import Image
        from io import BytesIO

        img = Image.new('RGB', (1, 1), color='red')
        buf = BytesIO()
        img.save(buf, format='JPEG')
        jpeg_data = buf.getvalue()

        images = [
            SimpleUploadedFile(f"img{i}.jpg", jpeg_data, content_type="image/jpeg")
            for i in range(6)
        ]
        payload = {
            "nomi": "Dazmol",
            "olchov_birligi": "dona",
            "kelish_narxi": "150000.00",
            "ustama": "20.00",
            "rasm": images,
            "qoldiqlar": '[{"dokon": %d, "miqdori": 10, "ogohlantirish": 2}]' % self.dokon.id
        }
        response = self.client.post(self.list_url, payload, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('rasm', response.data['errors'])


class MahsulotRasmAPITestCase(APITestCase):
    def setUp(self):
        self.image_list_url = reverse('product-image-list')
        self.biznes = Biznes.objects.create(nomi="Test Biznes", egasi_ism="Owner")
        self.manager_xodim = Xodim.objects.create(
            biznes=self.biznes,
            ism="Manager", familiya="Testov", telefon_raqam="+998907777777",
            parol="managerpassword123", rol="omborchi", jinsi="erkak"
        )
        self.manager_token = Token.objects.create(user=self.manager_xodim.user).key
        self.product = Mahsulot.objects.create(
            biznes=self.biznes,
            nomi="Dazmol",
            olchov_birligi="dona",
            kelish_narxi=Decimal("150000.00"),
            ustama=Decimal("20.00"),
            miqdori=10
        )

    def test_upload_single_image_to_product(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.manager_token)
        from django.core.files.uploadedfile import SimpleUploadedFile
        from PIL import Image
        from io import BytesIO

        img = Image.new('RGB', (1, 1), color='red')
        buf = BytesIO()
        img.save(buf, format='JPEG')
        jpeg_data = buf.getvalue()

        file = SimpleUploadedFile("single.jpg", jpeg_data, content_type="image/jpeg")
        payload = {
            "mahsulot": self.product.id,
            "rasm": file
        }
        response = self.client.post(self.image_list_url, payload, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(self.product.rasmlar.count(), 1)

    def test_upload_image_exceeding_limit(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.manager_token)
        from django.core.files.uploadedfile import SimpleUploadedFile
        from PIL import Image
        from io import BytesIO

        img = Image.new('RGB', (1, 1), color='red')
        buf = BytesIO()
        img.save(buf, format='JPEG')
        jpeg_data = buf.getvalue()

        # Create 5 images first
        for i in range(5):
            file = SimpleUploadedFile(f"file_{i}.jpg", jpeg_data, content_type="image/jpeg")
            MahsulotRasm.objects.create(mahsulot=self.product, rasm=file)

        # Uploading 6th image should fail
        file = SimpleUploadedFile("sixth.jpg", jpeg_data, content_type="image/jpeg")
        payload = {
            "mahsulot": self.product.id,
            "rasm": file
        }
        response = self.client.post(self.image_list_url, payload, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)


class ImportAPITestCase(APITestCase):
    def setUp(self):
        self.import_list_url = reverse('import-list')
        self.biznes = Biznes.objects.create(nomi="Test Biznes", egasi_ism="Owner")
        self.dokon = Dokon.objects.create(biznes=self.biznes, nomi="Test Dokon")

        self.admin_xodim = Xodim.objects.create(
            biznes=self.biznes,
            ism="Admin", familiya="Testov", telefon_raqam="+998909999999",
            parol="adminpassword123", rol="admin", jinsi="erkak"
        )
        self.admin_token = Token.objects.create(user=self.admin_xodim.user).key
        
        # Create cashier (restricted role)
        self.cashier_xodim = Xodim.objects.create(
            biznes=self.biznes,
            ism="Cashier", familiya="Restricted", telefon_raqam="+998908888888",
            parol="cashierpassword123", rol="sotuvchi", jinsi="ayol"
        )
        self.cashier_token = Token.objects.create(user=self.cashier_xodim.user).key

    def test_import_csv_flow(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.admin_token)
        from django.core.files.uploadedfile import SimpleUploadedFile
        csv_data = (
            "Nomi,Shtrix-kod,Miqdori,Kelish narxi,Sotish narxi,O'lchov birligi,Xususiyatlar\n"
            "Armatura,9948493123123,300,40000.00,60000.00,Штука,\"Brend: Quvasoy, Qalinligi: 1.5mm\"\n"
        )
        mock_file = SimpleUploadedFile("armatura.csv", csv_data.encode('utf-8'), content_type="text/csv")
        
        payload = {
            "nomi": "Armatura kirim",
            "fayl": mock_file,
            "import_turi": "kirim",
            "dokon": self.dokon.id
        }
        
        response = self.client.post(self.import_list_url, payload, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data['holat'], 'yakunlangan')
        self.assertEqual(response.data['miqdori'], 300)
        self.assertEqual(float(response.data['kelish_summasi']), 12000000.0)
        
        product = Mahsulot.objects.get(shtrix_kodlar__kod="9948493123123")
        self.assertEqual(product.nomi, "Armatura")
        self.assertEqual(product.characteristics.count(), 2)
        self.assertTrue(product.characteristics.filter(name="Brend", value="Quvasoy").exists())
        self.assertTrue(product.characteristics.filter(name="Qalinligi", value="1.5mm").exists())

    def test_import_manual_elements_with_mahsulot_id(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.admin_token)
        prod = Mahsulot.objects.create(biznes=self.biznes, nomi="Lom", kelish_narxi="10000.00", sotish_narxi="15000.00")
        payload = {
            "nomi": "Manual Kirim",
            "dokon": self.dokon.id,
            "import_turi": "kirim",
            "elementlar": [
                {
                    "mahsulot": prod.id,
                    "miqdori": "5.0",
                    "kelish_narx": 12000.00
                }
            ]
        }
        response = self.client.post(self.import_list_url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data['holat'], 'yakunlangan')
        self.assertEqual(response.data['miqdori'], 5)

    def test_import_invalid_file_extension(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.admin_token)
        from django.core.files.uploadedfile import SimpleUploadedFile
        # Text file is invalid
        mock_file = SimpleUploadedFile("invalid.txt", b"some plain text content", content_type="text/plain")
        
        payload = {
            "nomi": "Invalid Extension Import",
            "fayl": mock_file
        }
        response = self.client.post(self.import_list_url, payload, format='multipart')
        # Should raise bad request due to file extension validator
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_import_invalid_row_values(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.admin_token)
        from django.core.files.uploadedfile import SimpleUploadedFile
        # Row 2 contains selling price < cost price, and row 3 contains negative quantity
        csv_data = (
            "Nomi,Shtrix-kod,Miqdori,Kelish narxi,Sotish narxi,O'lchov birligi\n"
            "Armatura,9948493123123,300,50000.00,40000.00,Штука\n"
            "Gips,9948493124124,-10,10000.00,12000.00,кг\n"
        )
        mock_file = SimpleUploadedFile("bad_values.csv", csv_data.encode('utf-8'), content_type="text/csv")
        
        payload = {
            "nomi": "Bad Values Import",
            "fayl": mock_file
        }
        response = self.client.post(self.import_list_url, payload, format='multipart')
        print("IMPORT INVALID VALUES RESPONSE:", response.data)
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('fayl', response.data['errors'])
        # Verify row error description is returned
        self.assertTrue(any("Qator 2" in err or "Qator 3" in err for err in response.data['errors']['fayl']))

    def test_import_cashier_access_denied(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.cashier_token)
        from django.core.files.uploadedfile import SimpleUploadedFile
        csv_data = "Nomi,Shtrix-kod,Miqdori,Kelish narxi,Sotish narxi,O'lchov birligi\nArmatura,9948493123123,300,40000.00,60000.00,Штука\n"
        mock_file = SimpleUploadedFile("armatura.csv", csv_data.encode('utf-8'), content_type="text/csv")
        
        payload = {
            "nomi": "Cashier Try Import",
            "fayl": mock_file
        }
        response = self.client.post(self.import_list_url, payload, format='multipart')
        # Cashier should get 403 Forbidden
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_import_manual_flow(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.admin_token)
        payload = {
            "nomi": "Manual import test",
            "import_turi": "kirim",
            "dokon": self.dokon.id,
            "elementlar": [
                {
                    "nomi": "Manual Cement",
                    "shtrix_kod": "888877776666",
                    "miqdori": 50,
                    "kelish_narxi": 20000.0,
                    "sotish_narxi": 30000.0,
                    "olchov_birligi": "kg",
                    "toifa": "Stroy",
                    "characteristics": [
                        {"name": "Qalinligi", "value": "2mm"}
                    ]
                }
            ]
        }
        response = self.client.post(self.import_list_url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data['holat'], 'yakunlangan')
        self.assertEqual(response.data['miqdori'], 50)
        self.assertEqual(float(response.data['kelish_summasi']), 1000000.0)

        # Verify product was created and stock added
        product = Mahsulot.objects.get(shtrix_kodlar__kod="888877776666")
        self.assertEqual(product.nomi, "Manual Cement")
        self.assertEqual(product.miqdori, 50)
        self.assertTrue(product.characteristics.filter(name="Qalinligi", value="2mm").exists())

    def test_import_return_flow(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.admin_token)
        from products.models import Taminotchi, DokonQoldiq
        supplier = Taminotchi.objects.create(biznes=self.biznes, nomi="Test Supplier")
        
        # Create product with 30 initial stock
        product = Mahsulot.objects.create(
            biznes=self.biznes, nomi="Brick return test",
            kelish_narxi="1000.00", sotish_narxi="1500.00", toifa="Material"
        )
        product._custom_barcodes = ["555544443333"]
        product.save()
        
        dq = DokonQoldiq.objects.create(mahsulot=product, dokon=self.dokon, miqdori=30)
        product.miqdori = 30
        product.save()

        payload = {
            "nomi": "Manual return test",
            "import_turi": "qaytarish",
            "dokon": self.dokon.id,
            "taminotchi": supplier.id,
            "tolov_turi": "nasiya",
            "elementlar": [
                {
                    "nomi": "Brick return test",
                    "shtrix_kod": "555544443333",
                    "miqdori": 10,
                    "kelish_narxi": 1000.0,
                    "sotish_narxi": 1500.0
                }
            ]
        }
        
        response = self.client.post(self.import_list_url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data['holat'], 'yakunlangan')
        self.assertEqual(response.data['import_turi'], 'qaytarish')
        self.assertEqual(response.data['tolov_turi'], 'nasiya')
        self.assertEqual(response.data['taminotchi'], supplier.id)

        # Verify stock decreased to 20
        dq.refresh_from_db()
        self.assertEqual(dq.miqdori, 20)
        
        product.refresh_from_db()
        self.assertEqual(product.miqdori, 20)

        # Verify list filtering
        filter_url = f"{self.import_list_url}?taminotchi={supplier.id}&tolov_turi=nasiya"
        list_response = self.client.get(filter_url)
        self.assertEqual(list_response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(list_response.data['results']), 1)
        self.assertEqual(list_response.data['results'][0]['id'], response.data['id'])

    def test_import_stats_and_column_mapping(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.admin_token)
        
        # 1. API level stats check
        payload1 = {
            "nomi": "Kirim 1",
            "import_turi": "kirim",
            "dokon": self.dokon.id,
            "elementlar": [
                {"nomi": "Cement A", "miqdori": 100, "kelish_narxi": 5000.0, "sotish_narxi": 7000.0}
            ]
        }
        res1 = self.client.post(self.import_list_url, payload1, format='json')
        self.assertEqual(res1.status_code, status.HTTP_201_CREATED)
        self.assertTrue(res1.data['chek_raqami'].startswith('#'))

        payload2 = {
            "nomi": "Kirim 2",
            "import_turi": "kirim",
            "dokon": self.dokon.id,
            "elementlar": [
                {"nomi": "Cement B", "miqdori": 50, "kelish_narxi": 4000.0, "sotish_narxi": 6000.0}
            ]
        }
        res2 = self.client.post(self.import_list_url, payload2, format='json')
        self.assertEqual(res2.status_code, status.HTTP_201_CREATED)

        stats_url = f"{self.import_list_url}stats/?dokon={self.dokon.id}"
        stats_res = self.client.get(stats_url)
        self.assertEqual(stats_res.status_code, status.HTTP_200_OK)
        # We have test_import_return_flow running too (adding 1 more import if db not cleared)
        # So we just verify stats_res.data contains keys:
        self.assertIn('cheklar', stats_res.data)
        self.assertIn('soni', stats_res.data)
        self.assertIn('jami', stats_res.data)
        self.assertGreaterEqual(int(stats_res.data['cheklar']), 2)

        # 2. Model level custom column mapping check
        from django.core.files.uploadedfile import SimpleUploadedFile
        from products.models import Import
        csv_data = (
            "Nomi_Custom,Shtrix_Custom,Miqdor_Custom,Kelish_Custom,Sotish_Custom,Olchov_Custom,Qalinligi_Custom\n"
            "Brick,1122334455,5,100.00,150.00,dona,2.5mm\n"
        )
        mock_file = SimpleUploadedFile("custom_columns.csv", csv_data.encode('utf-8'), content_type="text/csv")
        mapping = {
            "nomi": 0,
            "shtrix_kod": 1,
            "miqdori": 2,
            "kelish_narxi": 3,
            "sotish_narxi": 4,
            "olchov_birligi": 5,
            "Qalinligi_Custom": 6
        }
        
        import_obj = Import.objects.create(
            biznes=self.biznes,
            dokon=self.dokon,
            nomi="Model mapped import",
            fayl=mock_file,
            column_mapping=mapping
        )
        import_obj.parse_and_save_elements()
        
        elements = import_obj.elementlar
        self.assertEqual(len(elements), 1)
        self.assertEqual(elements[0]['nomi'], "Brick")
        self.assertEqual(elements[0]['shtrix_kod'], "1122334455")
        self.assertEqual(elements[0]['miqdori'], 5)
        self.assertEqual(elements[0]['characteristics'][0]['name'], "Qalinligi_Custom")
        self.assertEqual(elements[0]['characteristics'][0]['value'], "2.5mm")



from .models import Dokon, Transfer

class TransferAPITestCase(APITestCase):
    def setUp(self):
        self.dokon_list_url = reverse('dokon-list')
        self.transfer_list_url = reverse('transfer-list')

        # Create Biznes
        self.biznes = Biznes.objects.create(nomi="Test Biznes", egasi_ism="Owner")

        # Create Shops
        self.dokon_a = Dokon.objects.create(biznes=self.biznes, nomi="Shop A", tavsif="Sender Shop")
        self.dokon_b = Dokon.objects.create(biznes=self.biznes, nomi="Shop B", tavsif="Receiver Shop")

        # Create Products
        self.prod_x = Mahsulot.objects.create(
            biznes=self.biznes,
            nomi="Cement",
            shtrix_kod="12345678",
            olchov_birligi="kg",
            kelish_narxi=Decimal("10000.00"),
            ustama=Decimal("20.00"),
            miqdori=50,
            ogohlantirish=10
        )
        DokonQoldiq.objects.create(mahsulot=self.prod_x, dokon=self.dokon_a, miqdori=50, ogohlantirish=10)

        self.prod_y = Mahsulot.objects.create(
            biznes=self.biznes,
            nomi="Brick",
            shtrix_kod="87654321",
            olchov_birligi="dona",
            kelish_narxi=Decimal("2000.00"),
            ustama=Decimal("15.00"),
            miqdori=100,
            ogohlantirish=20
        )
        DokonQoldiq.objects.create(mahsulot=self.prod_y, dokon=self.dokon_a, miqdori=100, ogohlantirish=20)

        # Create Admin
        self.admin_xodim = Xodim.objects.create(
            biznes=self.biznes,
            ism="Admin", familiya="Testov", telefon_raqam="+998909999999",
            parol="adminpassword123", rol="admin", jinsi="erkak"
        )
        self.admin_token = Token.objects.create(user=self.admin_xodim.user).key

        # Create Cashier
        self.cashier_xodim = Xodim.objects.create(
            biznes=self.biznes,
            ism="Cashier", familiya="Restricted", telefon_raqam="+998908888888",
            parol="cashierpassword123", rol="sotuvchi", jinsi="ayol"
        )
        self.cashier_token = Token.objects.create(user=self.cashier_xodim.user).key

    def test_dokon_list_and_create_permissions(self):
        # Cashier can list shops
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.cashier_token)
        response = self.client.get(self.dokon_list_url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 2)

        # Cashier cannot create shop
        response = self.client.post(self.dokon_list_url, {"nomi": "New Shop"})
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

        # Admin can create shop
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.admin_token)
        response = self.client.post(self.dokon_list_url, {"nomi": "New Shop"})
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

    def test_transfer_via_json(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.admin_token)
        payload = {
            "nomi": "Transfer Cement",
            "dokondan": self.dokon_a.id,
            "dokonga": self.dokon_b.id,
            "elementlar": [
                {
                    "nomi": "Cement",
                    "shtrix_kod": "12345678",
                    "miqdori": 10
                }
            ]
        }
        response = self.client.post(self.transfer_list_url, payload, format='json')
        print("TRANSFER RESPONSE DATA:", response.data)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data['holat'], 'yakunlangan') # auto-confirmed
        self.assertEqual(response.data['miqdori'], 10)
        self.assertEqual(float(response.data['summa']), 120000.0)

        # Verify stock decreased in sender shop and increased in receiver shop
        q_sender = DokonQoldiq.objects.get(mahsulot=self.prod_x, dokon=self.dokon_a)
        self.assertEqual(q_sender.miqdori, 40)
        q_receiver = DokonQoldiq.objects.get(mahsulot=self.prod_x, dokon=self.dokon_b)
        self.assertEqual(q_receiver.miqdori, 10)

        # Test search by ID
        t_id = response.data['id']
        search_res = self.client.get(self.transfer_list_url, {"search": str(t_id)})
        self.assertEqual(search_res.status_code, status.HTTP_200_OK)
        ids = [item['id'] for item in search_res.data]
        self.assertIn(t_id, ids)

    def test_transfer_via_excel(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.admin_token)
        from django.core.files.uploadedfile import SimpleUploadedFile
        import openpyxl
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nomi", "Shtrix-kod", "Miqdori", "O'lchov birligi"])
        ws.append(["Brick", "87654321", "30", "dona"])

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        mock_file = SimpleUploadedFile("transfer.xlsx", excel_file.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        payload = {
            "nomi": "Excel Transfer",
            "dokondan": self.dokon_a.id,
            "dokonga": self.dokon_b.id,
            "fayl": mock_file
        }
        response = self.client.post(self.transfer_list_url, payload, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data['holat'], 'yakunlangan')
        self.assertEqual(response.data['miqdori'], 30)
        self.assertEqual(float(response.data['summa']), 69000.0)

        # Verify stock decreased in sender shop and increased in receiver shop
        q_sender = DokonQoldiq.objects.get(mahsulot=self.prod_y, dokon=self.dokon_a)
        self.assertEqual(q_sender.miqdori, 70)
        q_receiver = DokonQoldiq.objects.get(mahsulot=self.prod_y, dokon=self.dokon_b)
        self.assertEqual(q_receiver.miqdori, 30)

    def test_transfer_invalid_product(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.admin_token)
        payload = {
            "nomi": "Transfer Fake",
            "dokondan": self.dokon_a.id,
            "dokonga": self.dokon_b.id,
            "elementlar": [
                {
                    "nomi": "Fake Product",
                    "shtrix_kod": "99999999",
                    "miqdori": 10
                }
            ]
        }
        response = self.client.post(self.transfer_list_url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_transfer_insufficient_stock(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.admin_token)
        payload = {
            "nomi": "Over Transfer",
            "dokondan": self.dokon_a.id,
            "dokonga": self.dokon_b.id,
            "elementlar": [
                {
                    "nomi": "Cement",
                    "shtrix_kod": "12345678",
                    "miqdori": 60  # Only 50 in stock
                }
            ]
        }
        response = self.client.post(self.transfer_list_url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_transfer_same_shop(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.admin_token)
        payload = {
            "nomi": "Same Shop Transfer",
            "dokondan": self.dokon_a.id,
            "dokonga": self.dokon_a.id,
            "elementlar": [
                {
                    "nomi": "Cement",
                    "shtrix_kod": "12345678",
                    "miqdori": 5
                }
            ]
        }
        response = self.client.post(self.transfer_list_url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)


class MahsulotShtrixKodAPITestCase(APITestCase):
    def setUp(self):
        self.barcodes_url = reverse('product-barcode-list')
        self.biznes = Biznes.objects.create(nomi="Test Biznes", egasi_ism="Owner")
        self.manager_xodim = Xodim.objects.create(
            biznes=self.biznes,
            ism="Manager", familiya="Testov", telefon_raqam="+998907777777",
            parol="managerpassword123", rol="omborchi", jinsi="erkak"
        )
        self.manager_token = Token.objects.create(user=self.manager_xodim.user).key
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.manager_token)

        self.product = Mahsulot.objects.create(
            biznes=self.biznes,
            nomi="Cement",
            shtrix_kod="12345678",
            olchov_birligi="kg",
            kelish_narxi=Decimal("10000.00"),
            ustama=Decimal("20.00"),
            miqdori=50,
            ogohlantirish=10
        )

    def test_list_barcodes_for_product(self):
        # We already have one generated barcode in setUp
        response = self.client.get(self.barcodes_url, {'mahsulot': self.product.id})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]['kod'], "12345678")

    def test_add_barcode_success(self):
        payload = {
            "mahsulot": self.product.id,
            "kod": "9876543210123"
        }
        response = self.client.post(self.barcodes_url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(self.product.shtrix_kodlar.count(), 2)

    def test_add_barcode_validation(self):
        # Duplicate barcode
        payload = {
            "mahsulot": self.product.id,
            "kod": "12345678"
        }
        response = self.client.post(self.barcodes_url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('kod', response.data['errors'])

        # Invalid length
        payload = {
            "mahsulot": self.product.id,
            "kod": "12345"
        }
        response = self.client.post(self.barcodes_url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_delete_barcode_protection(self):
        # Trying to delete the only barcode should fail
        only_bc = self.product.shtrix_kodlar.first()
        delete_url = reverse('product-barcode-detail', kwargs={'pk': only_bc.id})
        response = self.client.delete(delete_url)
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('detail', response.data['errors'])

        # Add second barcode, then delete should work
        new_bc = MahsulotShtrixKod.objects.create(mahsulot=self.product, kod="9876543210123")
        response = self.client.delete(delete_url)
        self.assertEqual(response.status_code, status.HTTP_204_NO_CONTENT)
        self.assertEqual(self.product.shtrix_kodlar.count(), 1)


class WriteOffAPITestCase(APITestCase):
    def setUp(self):
        self.write_off_list_url = reverse('write-off-list')
        self.biznes1 = Biznes.objects.create(nomi="Biznes 1", egasi_ism="Owner 1")
        self.biznes2 = Biznes.objects.create(nomi="Biznes 2", egasi_ism="Owner 2")

        self.admin1 = Xodim.objects.create(
            biznes=self.biznes1,
            ism="AdminBir", familiya="Testov", telefon_raqam="+998901111111",
            parol="adminpassword123!", rol="admin", jinsi="erkak"
        )
        self.admin1_token = Token.objects.create(user=self.admin1.user).key

        self.admin2 = Xodim.objects.create(
            biznes=self.biznes2,
            ism="AdminIkki", familiya="Testov", telefon_raqam="+998902222222",
            parol="adminpassword234!", rol="admin", jinsi="erkak"
        )
        self.admin2_token = Token.objects.create(user=self.admin2.user).key

        self.dokon1 = Dokon.objects.create(biznes=self.biznes1, nomi="Store 1")
        self.dokon2 = Dokon.objects.create(biznes=self.biznes2, nomi="Store 2")

        self.prod_a = Mahsulot.objects.create(
            biznes=self.biznes1, nomi="Cement A", olchov_birligi="kg",
            kelish_narxi=Decimal("10000.00"), sotish_narxi=Decimal("15000.00"), miqdori=100
        )
        MahsulotShtrixKod.objects.create(mahsulot=self.prod_a, kod="11111111")
        self.qoldiq_a = DokonQoldiq.objects.create(mahsulot=self.prod_a, dokon=self.dokon1, miqdori=100)

        self.prod_b = Mahsulot.objects.create(
            biznes=self.biznes2, nomi="Cement B", olchov_birligi="kg",
            kelish_narxi=Decimal("10000.00"), sotish_narxi=Decimal("15000.00"), miqdori=100
        )
        MahsulotShtrixKod.objects.create(mahsulot=self.prod_b, kod="22222222")
        self.qoldiq_b = DokonQoldiq.objects.create(mahsulot=self.prod_b, dokon=self.dokon2, miqdori=100)

    def test_write_off_manual_workflow(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.admin1_token)

        # 1. Create a draft write-off
        payload = {
            "nomi": "Yaroqsiz Mahsulotlar",
            "dokon": self.dokon1.id,
            "sababi": "defekt",
            "fayldan_hisobdan_chiqarish": False,
            "elementlar": [
                {
                    "mahsulot": self.prod_a.id,
                    "miqdori": 10
                }
            ]
        }
        response = self.client.post(self.write_off_list_url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        write_off_id = response.data['id']
        self.assertEqual(response.data['holat'], 'qoralama')
        self.assertEqual(response.data['miqdori'], 10)
        self.assertEqual(float(response.data['kelish_summasi']), 100000.0) # 10 * 10000
        self.assertEqual(float(response.data['sotish_summasi']), 150000.0) # 10 * 15000

        # 2. Try to confirm and execute
        confirm_url = reverse('write-off-confirm', kwargs={'pk': write_off_id})
        response = self.client.post(confirm_url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['holat'], 'yakunlangan')

        # Check stock is deducted
        self.qoldiq_a.refresh_from_db()
        self.assertEqual(self.qoldiq_a.miqdori, 90) # 100 - 10
        self.prod_a.refresh_from_db()
        self.assertEqual(self.prod_a.miqdori, 90)

    def test_write_off_insufficient_stock(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.admin1_token)

        payload = {
            "nomi": "Excess Write-off",
            "dokon": self.dokon1.id,
            "sababi": "defekt",
            "fayldan_hisobdan_chiqarish": False,
            "elementlar": [
                {
                    "mahsulot": self.prod_a.id,
                    "miqdori": 150 # Exceeds 100
                }
            ]
        }
        response = self.client.post(self.write_off_list_url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        write_off_id = response.data['id']

        # Confirm should fail
        confirm_url = reverse('write-off-confirm', kwargs={'pk': write_off_id})
        response = self.client.post(confirm_url)
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_write_off_excel_import(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.admin1_token)
        from django.core.files.uploadedfile import SimpleUploadedFile
        import openpyxl
        from io import BytesIO

        # Create template workbook in memory
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nomi", "Shtrix-kod", "Miqdori", "Kelish narxi"])
        ws.append(["Cement A", "11111111", "20", "10000.00"])

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        mock_file = SimpleUploadedFile("write_off.xlsx", excel_file.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        payload = {
            "nomi": "Excel Write-off",
            "dokon": self.dokon1.id,
            "sababi": "yoqotish",
            "fayldan_hisobdan_chiqarish": True,
            "fayl": mock_file
        }
        response = self.client.post(self.write_off_list_url, payload, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data['miqdori'], 20)
        self.assertEqual(float(response.data['kelish_summasi']), 200000.0)

    def test_write_off_excel_missing_file(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.admin1_token)
        payload = {
            "nomi": "Missing Excel File",
            "dokon": self.dokon1.id,
            "sababi": "yoqotish",
            "fayldan_hisobdan_chiqarish": True
            # No file attached
        }
        response = self.client.post(self.write_off_list_url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('fayl', response.data['errors'])

    def test_write_off_saas_tenant_isolation(self):
        # Admin 2 tries to access Admin 1's store in payload
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.admin2_token)
        payload = {
            "nomi": "Illegal Store",
            "dokon": self.dokon1.id, # Belongs to business 1
            "sababi": "yoqotish",
            "fayldan_hisobdan_chiqarish": False,
            "elementlar": [
                {
                    "mahsulot": self.prod_b.id,
                    "miqdori": 5
                }
            ]
        }
        response = self.client.post(self.write_off_list_url, payload, format='json')
        # Should raise validation error as store does not belong to business 2
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('dokon', response.data['errors'])


class ToplamAPITestCase(APITestCase):
    def setUp(self):
        from user.models import Biznes, Xodim, Tarif
        from products.models import Dokon, Mahsulot, DokonQoldiq
        from rest_framework.authtoken.models import Token
        from django.contrib.auth.models import User

        # Create tarif
        self.tarif = Tarif.objects.create(nomi="Pro", dokon_limiti=5, mahsulot_limiti=100, xodim_limiti=5)
        
        # Create businesses
        self.biznes1 = Biznes.objects.create(nomi="Biznes 1", egasi_ism="Owner 1", tarif=self.tarif)
        self.biznes2 = Biznes.objects.create(nomi="Biznes 2", egasi_ism="Owner 2", tarif=self.tarif)

        # Create users
        self.u1 = User.objects.create_user(username="u1", password="p1")
        self.u2 = User.objects.create_user(username="u2", password="p2")

        # Create xodims
        self.x1 = Xodim.objects.create(user=self.u1, ism="Ali", familiya="Valiyev", telefon_raqam="+998901234567", parol="secret123", jinsi="erkak", biznes=self.biznes1, rol="admin")
        self.x2 = Xodim.objects.create(user=self.u2, ism="Bobur", familiya="Karimov", telefon_raqam="+998901234568", parol="secret123", jinsi="erkak", biznes=self.biznes2, rol="admin")

        # Create tokens
        self.t1 = Token.objects.create(user=self.u1).key
        self.t2 = Token.objects.create(user=self.u2).key

        # Create stores
        self.dokon1 = Dokon.objects.create(biznes=self.biznes1, nomi="Store 1")
        self.dokon2 = Dokon.objects.create(biznes=self.biznes2, nomi="Store 2")

        # Create products
        self.p1 = Mahsulot.objects.create(biznes=self.biznes1, nomi="Cement", olchov_birligi="dona", kelish_narxi=Decimal("100.00"), sotish_narxi=Decimal("150.00"), miqdori=0)
        self.p2 = Mahsulot.objects.create(biznes=self.biznes2, nomi="Metal", olchov_birligi="dona", kelish_narxi=Decimal("200.00"), sotish_narxi=Decimal("300.00"), miqdori=0)

        # Create base stock
        self.dq1 = DokonQoldiq.objects.create(mahsulot=self.p1, dokon=self.dokon1, miqdori=10, ogohlantirish=0)
        self.p1.miqdori = 10
        self.p1.save()

        # URLs
        self.list_url = reverse('toplam-list')

    def test_create_toplam_replenishment_success(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.t1)
        payload = {
            "nomi": "Topilgan batch",
            "dokon": self.dokon1.id,
            "elementlar": [
                {
                    "mahsulot": self.p1.id,
                    "miqdori": 15
                }
            ]
        }
        response = self.client.post(self.list_url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data['holat'], 'yakunlangan') # auto-confirmed

        # Verify stock levels
        self.dq1.refresh_from_db()
        self.assertEqual(self.dq1.miqdori, 25) # 10 initial + 15 added

        self.p1.refresh_from_db()
        self.assertEqual(self.p1.miqdori, 25)

    def test_toplam_saas_isolation(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.t2)
        payload = {
            "nomi": "Illegal replenishment",
            "dokon": self.dokon1.id, # belongs to business 1
            "elementlar": [
                {
                    "mahsulot": self.p2.id,
                    "miqdori": 5
                }
            ]
        }
        response = self.client.post(self.list_url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('dokon', response.data['errors'])

    def test_create_composite_bundle_success(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.t1)
        payload = {
            "nomi": "Cement-Metal Super Bundle",
            "dokon": self.dokon1.id,
            "miqdori": 5,
            "shtrix_kod": "121212121212",
            "kelish_narxi": 400.00,
            "sotish_narxi": 500.00,
            "characteristics": [
                {"name": "Rang", "value": "Grey"}
            ],
            "elementlar": [
                {
                    "mahsulot": self.p1.id,
                    "miqdori": 2
                }
            ]
        }
        # First, ensure confirm_and_execute can run by executing perform_create which confirms
        response = self.client.post(self.list_url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data['holat'], 'yakunlangan')

        # Verify component stock level (dq1 started with 10, consumed 5 * 2 = 10, so should be 0)
        self.dq1.refresh_from_db()
        self.assertEqual(self.dq1.miqdori, 0)

        # Verify bundle product is created and has stock of 5
        bundle_prod = Mahsulot.objects.get(shtrix_kodlar__kod="121212121212")
        self.assertEqual(bundle_prod.nomi, "Cement-Metal Super Bundle")
        self.assertEqual(bundle_prod.miqdori, 5)
        self.assertTrue(bundle_prod.characteristics.filter(name="Rang", value="Grey").exists())


class ImportAPITestCase(APITestCase):
    def setUp(self):
        from user.models import Biznes, Xodim, Tarif
        from products.models import Dokon, Mahsulot, XususiyatMaydoni
        from rest_framework.authtoken.models import Token
        from django.contrib.auth.models import User

        self.tarif = Tarif.objects.create(nomi="Pro", dokon_limiti=5, mahsulot_limiti=100, xodim_limiti=5)
        self.biznes = Biznes.objects.create(nomi="Import Biznes", egasi_ism="Importer", tarif=self.tarif)
        self.u1 = User.objects.create_user(username="importer_u", password="password123")
        self.x1 = Xodim.objects.create(
            user=self.u1, ism="Ali", familiya="Valiyev", telefon_raqam="+998909999999", 
            parol="secret123", jinsi="erkak", biznes=self.biznes, rol="admin"
        )
        self.t1 = Token.objects.create(user=self.u1).key
        self.dokon = Dokon.objects.create(biznes=self.biznes, nomi="Main Store")

        # Create custom field "Qalinligi"
        self.field1 = XususiyatMaydoni.objects.create(biznes=self.biznes, nomi="Qalinligi", tur="matn", is_active=True)

        self.list_url = reverse('import-list')
        self.template_url = reverse('import-template')

    def test_download_template_contains_dynamic_fields(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.t1)
        response = self.client.get(self.template_url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        # Parse Excel response in memory
        import openpyxl
        from io import BytesIO
        wb = openpyxl.load_workbook(filename=BytesIO(response.content), data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        self.assertIn("Nomi", headers)
        self.assertIn("Shtrix-kod", headers)
        self.assertIn("Qalinligi", headers) # Custom field included
        self.assertNotIn("Artikul", headers) # Artikul removed

    def test_import_execution_success(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.t1)
        
        # Create Excel import file
        import openpyxl
        from io import BytesIO
        from django.core.files.uploadedfile import SimpleUploadedFile
        
        wb = openpyxl.Workbook()
        ws = wb.active
        # Headers: Nomi, Shtrix-kod, Miqdori, Kelish narxi, Sotish narxi, Toifa, Brend, O'lchov birligi, Yetkazib beruvchi, Tavsif, Qalinligi
        ws.append([
            "Nomi", "Shtrix-kod", "Miqdori", "Kelish narxi", "Sotish narxi",
            "Toifa", "Brend", "O'lchov birligi", "Yetkazib beruvchi", "Tavsif", "Qalinligi"
        ])
        ws.append([
            "Imported Tovar", "9876543210123", "30", "5000.00", "7500.00",
            "Metallar", "Premium", "dona", "Premium supplier", "Great metal", "5mm"
        ])
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        fayl = SimpleUploadedFile("test_import.xlsx", excel_file.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        
        payload = {
            "nomi": "Test Excel Import",
            "dokon": self.dokon.id,
            "fayl": fayl,
            "import_turi": "kirim",
            "shtrixkod_generatsiya_qilish": False
        }
        
        response = self.client.post(self.list_url, payload, format='multipart')
        if response.status_code != 201:
            print("RESPONSE DATA:", response.data)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data['holat'], 'yakunlangan') # auto-confirmed
        
        # Verify product created
        from products.models import Mahsulot, DokonQoldiq, Taminotchi
        product = Mahsulot.objects.get(biznes=self.biznes, nomi="Imported Tovar")
        self.assertEqual(product.kelish_narxi, Decimal("5000.00"))
        self.assertEqual(product.sotish_narxi, Decimal("7500.00"))
        self.assertEqual(product.toifa, "Metallar")
        self.assertEqual(product.brend, "Premium")
        self.assertEqual(product.tavsif, "Great metal")
        
        # Check supplier
        self.assertEqual(product.taminotchi.nomi, "Premium supplier")
        self.assertEqual(product.taminotchi.biznes, self.biznes)
        
        # Check stock level
        dq = DokonQoldiq.objects.get(mahsulot=product, dokon=self.dokon)
        self.assertEqual(dq.miqdori, 30)
        self.assertEqual(product.miqdori, 30)
        
        # Check characteristic
        char = product.characteristics.filter(name="Qalinligi", value="5mm").first()
        self.assertIsNotNone(char)


class BulkOperationsAPITestCase(APITestCase):
    def setUp(self):
        from user.models import Biznes, Xodim, Tarif
        from products.models import Dokon, Mahsulot
        from rest_framework.authtoken.models import Token
        from django.contrib.auth.models import User

        self.tarif = Tarif.objects.create(nomi="Pro", dokon_limiti=5, mahsulot_limiti=100, xodim_limiti=5)
        self.biznes1 = Biznes.objects.create(nomi="Biznes 1", egasi_ism="Owner 1", tarif=self.tarif)
        self.biznes2 = Biznes.objects.create(nomi="Biznes 2", egasi_ism="Owner 2", tarif=self.tarif)

        self.u1 = User.objects.create_user(username="u1_bulk", password="password123")
        self.u2 = User.objects.create_user(username="u2_bulk", password="password123")

        self.x1 = Xodim.objects.create(
            user=self.u1, ism="Ali", familiya="Valiyev", telefon_raqam="+998901111111", 
            parol="secret123", jinsi="erkak", biznes=self.biznes1, rol="admin"
        )
        self.x2 = Xodim.objects.create(
            user=self.u2, ism="Bobur", familiya="Karimov", telefon_raqam="+998902222222", 
            parol="secret123", jinsi="erkak", biznes=self.biznes2, rol="admin"
        )

        self.t1 = Token.objects.create(user=self.u1).key
        self.t2 = Token.objects.create(user=self.u2).key

        self.p1 = Mahsulot.objects.create(biznes=self.biznes1, nomi="Metal A", olchov_birligi="dona", kelish_narxi=Decimal("1000.00"), sotish_narxi=Decimal("1500.00"), miqdori=10)
        self.p2 = Mahsulot.objects.create(biznes=self.biznes1, nomi="Metal B", olchov_birligi="dona", kelish_narxi=Decimal("2000.00"), sotish_narxi=Decimal("3000.00"), miqdori=20)
        self.p3 = Mahsulot.objects.create(biznes=self.biznes2, nomi="Cement C", olchov_birligi="kg", kelish_narxi=Decimal("1500.00"), sotish_narxi=Decimal("2000.00"), miqdori=15)

        self.bulk_url = reverse('mahsulot-bulk-operations')

    def test_bulk_archive(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.t1)
        payload = {
            "action": "archive",
            "product_ids": [self.p1.id, self.p2.id],
            "params": {"archive": True}
        }
        response = self.client.post(self.bulk_url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        self.p1.refresh_from_db()
        self.p2.refresh_from_db()
        self.assertFalse(self.p1.is_active)
        self.assertFalse(self.p2.is_active)

    def test_bulk_archive_saas_isolation(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.t2)
        payload = {
            "action": "archive",
            "product_ids": [self.p1.id],
            "params": {"archive": True}
        }
        response = self.client.post(self.bulk_url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    def test_bulk_set_low_stock(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.t1)
        payload = {
            "action": "set_low_stock",
            "product_ids": [self.p1.id, self.p2.id],
            "params": {"threshold": 12}
        }
        response = self.client.post(self.bulk_url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        self.p1.refresh_from_db()
        self.p2.refresh_from_db()
        self.assertEqual(self.p1.ogohlantirish, 12)
        self.assertEqual(self.p2.ogohlantirish, 12)

    def test_bulk_edit_prices_percentage_increase(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.t1)
        payload = {
            "action": "edit_prices",
            "product_ids": [self.p1.id, self.p2.id],
            "params": {
                "price_type": "sotish_narxi",
                "operation": "oshirish_foiz",
                "value": 10.00
            }
        }
        response = self.client.post(self.bulk_url, payload, format='json')
        if response.status_code != 200:
            print("BULK PRICES RESPONSE:", response.data)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        self.p1.refresh_from_db()
        self.p2.refresh_from_db()
        self.assertEqual(self.p1.sotish_narxi, Decimal("1650.00"))
        self.assertEqual(self.p2.sotish_narxi, Decimal("3300.00"))

    def test_bulk_edit_prices_erkin_narx(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.t1)
        payload = {
            "action": "edit_prices",
            "product_ids": [self.p1.id, self.p2.id],
            "params": {
                "erkin_narx": True
            }
        }
        response = self.client.post(self.bulk_url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        self.p1.refresh_from_db()
        self.p2.refresh_from_db()
        self.assertTrue(self.p1.erkin_narx)
        self.assertTrue(self.p2.erkin_narx)

    def test_bulk_edit_characteristics(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.t1)
        payload = {
            "action": "edit_characteristics",
            "product_ids": [self.p1.id, self.p2.id],
            "params": {
                "characteristics": {
                    "Qalinligi": "8mm",
                    "Rangi": "Ko'k"
                }
            }
        }
        response = self.client.post(self.bulk_url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        self.p1.refresh_from_db()
        self.p2.refresh_from_db()
        
        c1 = self.p1.characteristics.filter(name="Qalinligi", value="8mm").first()
        self.assertIsNotNone(c1)
        c2 = self.p2.characteristics.filter(name="Rangi", value="Ko'k").first()
        self.assertIsNotNone(c2)

    def test_bulk_print_labels(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.t1)
        payload = {
            "action": "print_labels",
            "product_ids": [self.p1.id, self.p2.id]
        }
        response = self.client.post(self.bulk_url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("labels", response.data)
        self.assertEqual(len(response.data["labels"]), 2)
        self.assertEqual(response.data["labels"][0]["nomi"], "Metal A")

    def test_bulk_print_labels_with_stock_quantity(self):
        from products.models import DokonQoldiq, Dokon
        dokon = Dokon.objects.create(biznes=self.biznes1, nomi="Test Shop")
        DokonQoldiq.objects.create(mahsulot=self.p1, dokon=dokon, miqdori=7)

        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.t1)
        
        payload = {
            "action": "print_labels",
            "product_ids": [self.p1.id],
            "params": {
                "soni_turi": "qoldiqlar_boyicha"
            }
        }
        response = self.client.post(self.bulk_url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["labels"][0]["soni"], 10) # global stock is 10

        payload_store = {
            "action": "print_labels",
            "product_ids": [self.p1.id],
            "params": {
                "soni_turi": "qoldiqlar_boyicha",
                "dokon": dokon.id
            }
        }
        response = self.client.post(self.bulk_url, payload_store, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["labels"][0]["soni"], 7) # store stock is 7

        # Zero stock skip verification
        from products.models import Mahsulot
        p_zero = Mahsulot.objects.create(biznes=self.biznes1, nomi="Zero Stock", olchov_birligi="dona", kelish_narxi=Decimal("100.00"), sotish_narxi=Decimal("150.00"), miqdori=0)

        payload_skip = {
            "action": "print_labels",
            "product_ids": [self.p1.id, p_zero.id],
            "params": {
                "nol_qoldiq_otkazish": True
            }
        }
        response = self.client.post(self.bulk_url, payload_skip, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data["labels"]), 1)
        self.assertEqual(response.data["labels"][0]["nomi"], "Metal A")


class YorliqShablonAPITestCase(APITestCase):
    def setUp(self):
        from user.models import Biznes, Xodim, Tarif
        from rest_framework.authtoken.models import Token
        from django.contrib.auth.models import User

        self.tarif = Tarif.objects.create(nomi="Pro", dokon_limiti=5, mahsulot_limiti=100, xodim_limiti=5)
        self.biznes1 = Biznes.objects.create(nomi="Biznes 1", egasi_ism="Owner 1", tarif=self.tarif)
        self.biznes2 = Biznes.objects.create(nomi="Biznes 2", egasi_ism="Owner 2", tarif=self.tarif)

        self.u1 = User.objects.create_user(username="u1_shablon", password="password123")
        self.u2 = User.objects.create_user(username="u2_shablon", password="password123")

        self.x1 = Xodim.objects.create(
            user=self.u1, ism="Ali", familiya="Valiyev", telefon_raqam="+998904444444", 
            parol="secret123", jinsi="erkak", biznes=self.biznes1, rol="admin"
        )
        self.x2 = Xodim.objects.create(
            user=self.u2, ism="Bobur", familiya="Karimov", telefon_raqam="+998905555555", 
            parol="secret123", jinsi="erkak", biznes=self.biznes2, rol="admin"
        )

        self.t1 = Token.objects.create(user=self.u1).key
        self.t2 = Token.objects.create(user=self.u2).key

        self.list_url = reverse('price-tag-templates-list')

    def test_create_and_read_template(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.t1)
        payload = {
            "nomi": "Standart 40x20",
            "eni": 40.0,
            "uzunlik": 20.0,
            "shtrixkod_formati": "CODE128",
            "xususiyatlar": ["nomi", "sotish_narxi", "shtrix_kod"]
        }
        response = self.client.post(self.list_url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data["nomi"], "Standart 40x20")

        response = self.client.get(self.list_url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]["xususiyatlar"], ["nomi", "sotish_narxi", "shtrix_kod"])

    def test_template_saas_isolation(self):
        from products.models import YorliqShablon
        shablon = YorliqShablon.objects.create(biznes=self.biznes1, nomi="Biznes 1 template", eni=40, uzunlik=20)

        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.t2)
        response = self.client.get(self.list_url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 0)

        detail_url = reverse('price-tag-templates-detail', kwargs={'pk': shablon.id})
        response = self.client.put(detail_url, {"nomi": "Hacked"}, format='json')
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)


class WriteOffAPITestCase(APITestCase):
    def setUp(self):
        from user.models import Biznes, Xodim, Tarif
        from products.models import Dokon, Mahsulot, WriteOff
        from rest_framework.authtoken.models import Token
        from django.contrib.auth.models import User

        self.tarif = Tarif.objects.create(nomi="Pro", dokon_limiti=5, mahsulot_limiti=100, xodim_limiti=5)
        self.biznes = Biznes.objects.create(nomi="Biznes 1", egasi_ism="Owner 1", tarif=self.tarif)

        self.u1 = User.objects.create_user(username="u1_writeoff", password="password123")
        self.x1 = Xodim.objects.create(
            user=self.u1, ism="Ali", familiya="Valiyev", telefon_raqam="+998906666666", 
            parol="secret123", jinsi="erkak", biznes=self.biznes, rol="admin"
        )
        self.t1 = Token.objects.create(user=self.u1).key

        self.dokon = Dokon.objects.create(biznes=self.biznes, nomi="WriteOff Shop")
        
        self.p1 = Mahsulot.objects.create(biznes=self.biznes, nomi="Cement A", olchov_birligi="kg", kelish_narxi=Decimal("100.00"), sotish_narxi=Decimal("150.00"), miqdori=10)

        # Create stock level
        from products.models import DokonQoldiq
        DokonQoldiq.objects.create(mahsulot=self.p1, dokon=self.dokon, miqdori=10)

        self.w1 = WriteOff.objects.create(biznes=self.biznes, dokon=self.dokon, nomi="Defect writeoff", sababi="defekt", holat="qoralama")
        
        self.list_url = reverse('write-off-list')

    def test_write_off_list_search_and_filter(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.t1)

        # Search by ID
        response = self.client.get(self.list_url, {"search": str(self.w1.id)})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 1)

        # Search by name
        response = self.client.get(self.list_url, {"search": "Defect"})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 1)

        # Filter by reason
        response = self.client.get(self.list_url, {"sababi": "defekt"})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 1)

        # Filter by reason (non-matching)
        response = self.client.get(self.list_url, {"sababi": "inventarizatsiya"})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 0)

        # Filter by date (sana)
        import datetime
        today_str = datetime.date.today().isoformat()
        response = self.client.get(self.list_url, {"sana": today_str})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 1)


class TaminotchiAPITestCase(APITestCase):
    def setUp(self):
        from user.models import Biznes, Xodim, Tarif
        from products.models import Dokon, Mahsulot, Taminotchi
        from orders.models import SupplierOrder, SupplierOrderItem
        from rest_framework.authtoken.models import Token
        from django.contrib.auth.models import User
        import datetime

        self.tarif = Tarif.objects.create(nomi="Pro", dokon_limiti=5, mahsulot_limiti=100, xodim_limiti=5)
        
        self.biznes1 = Biznes.objects.create(nomi="Biznes 1", egasi_ism="Owner 1", tarif=self.tarif)
        self.biznes2 = Biznes.objects.create(nomi="Biznes 2", egasi_ism="Owner 2", tarif=self.tarif)

        self.u1 = User.objects.create_user(username="u1_taminotchi", password="password123")
        self.x1 = Xodim.objects.create(
            user=self.u1, ism="Ali", familiya="Valiyev", telefon_raqam="+998906666666", 
            parol="secret123", jinsi="erkak", biznes=self.biznes1, rol="admin"
        )
        self.t1 = Token.objects.create(user=self.u1).key

        self.u2 = User.objects.create_user(username="u2_taminotchi", password="password123")
        self.x2 = Xodim.objects.create(
            user=self.u2, ism="Bobur", familiya="Sodiqov", telefon_raqam="+998907777777", 
            parol="secret123", jinsi="erkak", biznes=self.biznes2, rol="admin"
        )
        self.t2 = Token.objects.create(user=self.u2).key

        self.dokon = Dokon.objects.create(biznes=self.biznes1, nomi="Shop 1")
        self.taminotchi = Taminotchi.objects.create(biznes=self.biznes1, nomi="Husniddin", telefon_raqam="+998996070500")

        self.p1 = Mahsulot.objects.create(biznes=self.biznes1, nomi="Cement A", olchov_birligi="kg", kelish_narxi=Decimal("100.00"), sotish_narxi=Decimal("150.00"), miqdori=10)

        # Create Supplier Order
        self.order = SupplierOrder.objects.create(
            biznes=self.biznes1,
            taminotchi=self.taminotchi,
            dokon=self.dokon,
            nomi="Buyurtma 1",
            qabul_qilish_sanasi=datetime.date.today(),
            umumiy_summa=Decimal("1000.00"),
            tolangan_summa=Decimal("600.00")
        )
        SupplierOrderItem.objects.create(
            order=self.order,
            mahsulot=self.p1,
            miqdori=10,
            kelish_narxi=Decimal("100.00"),
            sotish_narxi=Decimal("150.00")
        )

        self.list_url = reverse('suppliers-list')

    def test_supplier_statistics_and_crud(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.t1)

        # Create new supplier with all requisites and extra details
        payload = {
            "nomi": "Islomjon",
            "telefon_raqam": "+998901234567",
            "telefonlar": ["+998901234567", "+998931112233"],
            "standart_ustama": "30.00",
            "eslatma": "Test notes here",
            "yuridik_nomi": "Islomjon LLC",
            "yuridik_manzil": "Tashkent City",
            "mamlakat": "Uzbekistan",
            "pochta_indeksi": "100000",
            "bank_hisob_raqami": "20208000600000000001",
            "bank_nomi_filiali": "Aloqabank, Tashkent",
            "inn": "301234567",
            "mfo": "00440"
        }
        response = self.client.post(self.list_url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data["nomi"], "Islomjon")
        self.assertEqual(response.data["telefonlar"], ["+998901234567", "+998931112233"])
        self.assertEqual(Decimal(response.data["standart_ustama"]), Decimal("30.00"))
        self.assertEqual(response.data["eslatma"], "Test notes here")
        self.assertEqual(response.data["yuridik_nomi"], "Islomjon LLC")
        self.assertEqual(response.data["yuridik_manzil"], "Tashkent City")
        self.assertEqual(response.data["mamlakat"], "Uzbekistan")
        self.assertEqual(response.data["pochta_indeksi"], "100000")
        self.assertEqual(response.data["bank_hisob_raqami"], "20208000600000000001")
        self.assertEqual(response.data["bank_nomi_filiali"], "Aloqabank, Tashkent")
        self.assertEqual(response.data["inn"], "301234567")
        self.assertEqual(response.data["mfo"], "00440")

        # Get list & check calculations
        response = self.client.get(self.list_url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        # Should have 2 suppliers (Husniddin, Islomjon)
        self.assertEqual(len(response.data), 2)
        
        # Find Husniddin and verify calculations
        husniddin_data = [item for item in response.data if item["nomi"] == "Husniddin"][0]
        self.assertEqual(Decimal(husniddin_data["buyurtmalar_summasi"]), Decimal("1000.00"))
        self.assertEqual(Decimal(husniddin_data["tolovlar_summasi"]), Decimal("600.00"))
        self.assertEqual(Decimal(husniddin_data["qarz_summasi"]), Decimal("400.00"))
        self.assertEqual(husniddin_data["tovarlar_soni"], 10)

        # Search by ID, name, or phone
        response = self.client.get(self.list_url, {"search": "Husniddin"})
        self.assertEqual(len(response.data), 1)

        response = self.client.get(self.list_url, {"search": "+998996070500"})
        self.assertEqual(len(response.data), 1)

    def test_supplier_saas_isolation(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.t2)
        response = self.client.get(self.list_url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 0)

    def test_supplier_pay_fifo(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.t1)
        
        # Transition the order to 'rasmiylashtirilgan' state so it accumulates debt
        self.order.rasmiylashtirish()
        
        pay_url = reverse('suppliers-pay', kwargs={'pk': self.taminotchi.id})
        payload = {
            "amount": "500.00",
            "tolov_turi": "naqd"
        }
        response = self.client.post(pay_url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        self.taminotchi.refresh_from_db()
        self.assertEqual(self.taminotchi.balans, Decimal("100.00"))
        
        self.order.refresh_from_db()
        self.assertEqual(self.order.tolangan_summa, Decimal("1000.00"))
        self.assertEqual(self.order.nasiya_summa, Decimal("0.00"))

    def test_supplier_dashboard(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.t1)
        
        self.order.rasmiylashtirish()
        
        from orders.models import SupplierOrderReturn
        ret = SupplierOrderReturn.objects.create(
            biznes=self.biznes1,
            order=self.order,
            dokon=self.dokon,
            taminotchi=self.taminotchi,
            holat='kutilmoqda'
        )
        
        from products.models import DokonQoldiq
        DokonQoldiq.objects.create(
            mahsulot=self.p1,
            dokon=self.dokon,
            miqdori=10
        )
        
        from orders.models import SupplierOrderPayment
        SupplierOrderPayment.objects.create(
            order=self.order,
            tolangan_summa=Decimal("600.00"),
            tolov_turi="naqd",
            xodim=self.x1
        )
        
        from orders.models import SupplierOrderReturnItem
        SupplierOrderReturnItem.objects.create(
            return_obj=ret,
            mahsulot=self.p1,
            miqdori=5,
            kelish_narxi=Decimal("40.00")
        )
        
        dashboard_url = reverse('suppliers-dashboard', kwargs={'pk': self.taminotchi.id})
        response = self.client.get(dashboard_url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        self.assertEqual(Decimal(response.data["balans"]), Decimal("0.00"))
        self.assertEqual(response.data["tolanmagan_buyurtmalar_count"], 1)
        self.assertEqual(response.data["tolangan_buyurtmalar_count"], 0)
        self.assertEqual(Decimal(response.data["buyurtmalar_summasi"]), Decimal("1000.00"))
        self.assertEqual(Decimal(response.data["tolovlar_summasi"]), Decimal("600.00"))
        self.assertEqual(Decimal(response.data["qarz_summasi"]), Decimal("400.00"))
        self.assertEqual(response.data["buyurtma_qilingan_mahsulotlar"], 10)
        self.assertEqual(response.data["qabul_qilingan_mahsulotlar"], 0)
        self.assertEqual(response.data["buyurtmalar_tezligi"], 1)
        self.assertEqual(Decimal(response.data["qaytarish_summasi"]), Decimal("200.00"))
        self.assertEqual(Decimal(response.data["qaytarilgan_tolovlar_summasi"]), Decimal("0.00"))

        payments_url = reverse('suppliers-payments', kwargs={'pk': self.taminotchi.id})
        res_pay = self.client.get(payments_url)
        self.assertEqual(res_pay.status_code, status.HTTP_200_OK)
        self.assertEqual(len(res_pay.data), 1)
        self.assertEqual(Decimal(res_pay.data[0]["tolangan_summa"]), Decimal("600.00"))

    def test_supplier_global_stats(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.t1)
        
        self.order.rasmiylashtirish()
        
        stats_url = reverse('suppliers-stats')
        response = self.client.get(stats_url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["yetkazib_beruvchilar_soni"], 1)
        self.assertEqual(Decimal(response.data["umumiy_buyurtmalar_summasi"]), Decimal("1000.00"))
        self.assertEqual(Decimal(response.data["umumiy_tolovlar_summasi"]), Decimal("600.00"))
        self.assertEqual(Decimal(response.data["umumiy_qarz_summasi"]), Decimal("400.00"))

    def test_supplier_list_dynamic_pagination(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.t1)
        
        response = self.client.get(self.list_url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(isinstance(response.data, list))
        
        response = self.client.get(self.list_url, {"page_size": 5})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(isinstance(response.data, dict))
        self.assertEqual(response.data["count"], 1)
        self.assertEqual(len(response.data["results"]), 1)

    def test_supplier_excel_exports(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.t1)
        
        # Test suppliers list export
        response = self.client.get(self.list_url, {"export": "excel"})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        
        # Test supplier payments list export
        payments_url = reverse('suppliers-payments', kwargs={'pk': self.taminotchi.id})
        response = self.client.get(payments_url, {"export": "excel"})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # Test imports list export
        response = self.client.get(reverse('import-list'), {"export": "excel"})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # Test write-offs list export
        response = self.client.get(reverse('write-off-list'), {"export": "excel"})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    def test_import_and_writeoff_excel_templates(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.t1)
        
        # Test imports template download
        response = self.client.get(reverse('import-template'))
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        
        # Test write-offs template download
        response = self.client.get(reverse('write-off-template'))
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    def test_previously_ordered_filtering(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.t1)
        
        p_ordered = Mahsulot.objects.create(
            biznes=self.biznes1, nomi="Ordered Product", olchov_birligi="dona",
            kelish_narxi=Decimal("100.00"), sotish_narxi=Decimal("150.00"), miqdori=10
        )
        p_not_ordered = Mahsulot.objects.create(
            biznes=self.biznes1, nomi="Not Ordered Product", olchov_birligi="dona",
            kelish_narxi=Decimal("200.00"), sotish_narxi=Decimal("300.00"), miqdori=5
        )
        
        from orders.models import SupplierOrder, SupplierOrderItem
        order = SupplierOrder.objects.create(
            biznes=self.biznes1, taminotchi=self.taminotchi, dokon=self.dokon,
            nomi="PO 1", holat="qoralama", qabul_qilish_sanasi="2026-07-20"
        )
        SupplierOrderItem.objects.create(
            order=order, mahsulot=p_ordered, miqdori=5, kelish_narxi=Decimal("100.00"),
            sotish_narxi=Decimal("150.00")
        )
        
        response = self.client.get(reverse('mahsulot-list'), {"oldin_buyurtma_qilingan": "true"})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        product_names = [p["nomi"] for p in response.data]
        self.assertIn("Ordered Product", product_names)
        self.assertNotIn("Not Ordered Product", product_names)
        
        response = self.client.get(reverse('mahsulot-list'), {"oldin_buyurtma_qilingan": "true", "taminotchi": self.taminotchi.id})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        product_names = [p["nomi"] for p in response.data]
        self.assertIn("Ordered Product", product_names)
        
        response = self.client.get(reverse('mahsulot-list'), {"oldin_buyurtma_qilingan": "true", "taminotchi": 9999})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 0)

    def test_supplier_initial_debt(self):
        self.client.credentials(HTTP_AUTHORIZATION='Token ' + self.t1)
        
        # Test creating supplier with initial debt using snake_case
        payload = {
            "nomi": "Taminotchi Snake",
            "dastlabki_qarz": "500000.00"
        }
        response = self.client.post(self.list_url, payload, format='json')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(Decimal(response.data["qarz_summasi"]), Decimal("500000.00"))
        self.assertEqual(Decimal(response.data["jami_qarz"]), Decimal("500000.00"))

        # Test creating supplier with initial debt using camelCase
        payload_camel = {
            "nomi": "Taminotchi Camel",
            "dastlabkiQarz": "300000.00"
        }
        response = self.client.post(self.list_url, payload_camel, format='json')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(Decimal(response.data["qarz_summasi"]), Decimal("300000.00"))

        # Test creating supplier with initial debt using oxirgi_qarz
        payload_oxirgi = {
            "nomi": "Taminotchi Oxirgi",
            "oxirgi_qarz": "150000.00"
        }
        response = self.client.post(self.list_url, payload_oxirgi, format='json')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(Decimal(response.data["qarz_summasi"]), Decimal("150000.00"))

        # Test creating supplier with initial debt using jamiQarz with spaces
        payload_jami = {
            "nomi": "Taminotchi Jami",
            "jamiQarz": "250 000.00"
        }
        response = self.client.post(self.list_url, payload_jami, format='json')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(Decimal(response.data["qarz_summasi"]), Decimal("250000.00"))

        # Test creating supplier when one field is empty string and the other is populated
        payload_mixed = {
            "nomi": "Taminotchi Mixed",
            "oxirgi_qarz": "",
            "jamiQarz": "400000.00"
        }
        response = self.client.post(self.list_url, payload_mixed, format='json')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(Decimal(response.data["qarz_summasi"]), Decimal("400000.00"))



class CompleteSystemWorkflowTestCase(APITestCase):
    """
    Multi-stage end-to-end integration test verifying the entire business flow:
    Registration -> Login -> Supplier & Store Setup -> Product Catalog -> Purchasing Order ->
    Receiving Inventory -> Balances & Payments -> Write-off -> Verification.
    """

    def test_complete_business_flow(self):
        from user.models import Xodim, Biznes
        from products.models import Mahsulot, Dokon, DokonQoldiq, WriteOff
        from orders.models import Taminotchi, SupplierOrder, SupplierOrderItem, SupplierOrderPayment
        
        # STAGE 1: Register new business and user
        register_url = reverse('register')
        reg_payload = {
            "ism": "Bekzod",
            "telefon_raqam": "+998909876543",
            "parol": "123456",
            "parolni_tasdiqlash": "123456",
            "biznes_nomi": "Temir Invest"
        }
        reg_response = self.client.post(register_url, reg_payload, format='json')
        self.assertEqual(reg_response.status_code, status.HTTP_201_CREATED)
        token = reg_response.data['token']
        self.client.credentials(HTTP_AUTHORIZATION=f'Token {token}')

        # Verify default entities created automatically by register API
        xodim = Xodim.objects.get(telefon_raqam="+998909876543")
        biznes = xodim.biznes
        self.assertIsNotNone(biznes)
        
        # Verify a default Dokon was created automatically in the register flow
        self.assertTrue(Dokon.objects.filter(biznes=biznes).exists())
        dokon = Dokon.objects.filter(biznes=biznes).first()

        # STAGE 2: Add a new supplier (Taminotchi)
        supplier_url = reverse('suppliers-list')
        sup_payload = {
            "nomi": "Metal Snab",
            "telefon_raqam": "+998935555555"
        }
        sup_response = self.client.post(supplier_url, sup_payload, format='json')
        self.assertEqual(sup_response.status_code, status.HTTP_201_CREATED)
        supplier_id = sup_response.data['id']
        taminotchi = Taminotchi.objects.get(id=supplier_id)

        # STAGE 3: Create a new product (Mahsulot) in the catalog
        product_url = reverse('mahsulot-list')
        prod_payload = {
            "nomi": "Armatura A500C",
            "olchov_birligi": "metr",
            "kelish_narxi": "8000.00",
            "sotish_narxi": "12000.00",
            "miqdori": 0,
            "qoldiqlar": [
                {
                    "dokon": dokon.id,
                    "miqdori": 0,
                    "ogohlantirish": 0
                }
            ]
        }
        prod_response = self.client.post(product_url, prod_payload, format='json')
        self.assertEqual(prod_response.status_code, status.HTTP_201_CREATED)
        product_id = prod_response.data['id']
        mahsulot = Mahsulot.objects.get(id=product_id)

        # STAGE 4: Place a Supplier Order (Draft)
        order_url = reverse('supplier-order-list')
        order_payload = {
            "taminotchi": supplier_id,
            "dokon": dokon.id,
            "nomi": "Birinchi O'tkazma",
            "qabul_qilish_sanasi": "2026-07-20",
            "elementlar": [
                {
                    "mahsulot": product_id,
                    "miqdori": 100,
                    "kelish_narxi": "8000.00",
                    "sotish_narxi": "12000.00"
                }
            ]
        }
        order_response = self.client.post(order_url, order_payload, format='json')
        self.assertEqual(order_response.status_code, status.HTTP_201_CREATED)
        order_id = order_response.data['id']
        order = SupplierOrder.objects.get(id=order_id)
        self.assertEqual(order.holat, 'qoralama')

        # STAGE 5: Confirm / Rasmiylashtirish the Order
        confirm_url = reverse('supplier-order-confirm', kwargs={'pk': order_id})
        conf_response = self.client.post(confirm_url, format='json')
        self.assertEqual(conf_response.status_code, status.HTTP_200_OK)
        order.refresh_from_db()
        self.assertEqual(order.holat, 'rasmiylashtirilgan')

        # STAGE 6: Receive / Qabul qilish the inventory
        receive_url = reverse('supplier-order-receive', kwargs={'pk': order_id})
        rec_response = self.client.post(receive_url, {"apply_new_prices": True}, format='json')
        self.assertEqual(rec_response.status_code, status.HTTP_200_OK)
        order.refresh_from_db()
        self.assertEqual(order.holat, 'qabul_qilingan')

        # Verify stock updates
        mahsulot.refresh_from_db()
        self.assertEqual(mahsulot.miqdori, 100)
        qoldiq = DokonQoldiq.objects.get(mahsulot=mahsulot, dokon=dokon)
        self.assertEqual(qoldiq.miqdori, 100)

        # STAGE 7: Supplier Debt and Payment check
        taminotchi.refresh_from_db()
        self.assertEqual(order.nasiya_summa, Decimal('800000.00')) # 100 * 8000
        
        # Pay 300,000 UZS to the supplier order
        pay_url = reverse('supplier-order-pay', kwargs={'pk': order_id})
        pay_payload = {
            "amount": "300000.00",
            "tolov_turi": "naqd"
        }
        pay_response = self.client.post(pay_url, pay_payload, format='json')
        self.assertEqual(pay_response.status_code, status.HTTP_200_OK)
        order.refresh_from_db()
        self.assertEqual(order.tolangan_summa, Decimal('300000.00'))
        self.assertEqual(order.nasiya_summa, Decimal('500000.00'))

        # STAGE 8: Write-off 10 items due to defect
        write_off_list_url = reverse('write-off-list')
        wo_payload = {
            "nomi": "Defekt tovarlarni hisobdan chiqarish",
            "dokon": dokon.id,
            "sababi": "defekt",
            "elementlar": [
                {
                    "mahsulot": product_id,
                    "miqdori": 10,
                    "kelish_narxi": "8000.00"
                }
            ]
        }
        wo_response = self.client.post(write_off_list_url, wo_payload, format='json')
        self.assertEqual(wo_response.status_code, status.HTTP_201_CREATED)
        wo_id = wo_response.data['id']
        write_off = WriteOff.objects.get(id=wo_id)
        self.assertEqual(write_off.holat, 'qoralama')

        # Confirm the write-off to deduct stock
        wo_confirm_url = reverse('write-off-confirm', kwargs={'pk': wo_id})
        wo_conf_response = self.client.post(wo_confirm_url, format='json')
        self.assertEqual(wo_conf_response.status_code, status.HTTP_200_OK)

        # Verify final stocks (100 - 10 = 90)
        mahsulot.refresh_from_db()
        self.assertEqual(mahsulot.miqdori, 90)
        qoldiq.refresh_from_db()
        self.assertEqual(qoldiq.miqdori, 90)

    def test_ombor_stats_and_csv_export(self):
        register_url = reverse('register')
        reg_payload = {
            "ism": "OmborUser",
            "telefon_raqam": "+998901112233",
            "parol": "123456",
            "parolni_tasdiqlash": "123456",
            "biznes_nomi": "Ombor Biznes"
        }
        reg_response = self.client.post(register_url, reg_payload, format='json')
        self.assertEqual(reg_response.status_code, status.HTTP_201_CREATED)
        token = reg_response.data['token']
        self.client.credentials(HTTP_AUTHORIZATION=f'Token {token}')

        ombor_stats_url = reverse('mahsulot-ombor-stats')
        response = self.client.get(ombor_stats_url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn('barchasi', response.data)
        self.assertIn('faol', response.data)
        self.assertIn('kam_qoldiq', response.data)

        # Test CSV Export
        list_url = reverse('mahsulot-list')
        csv_response = self.client.get(list_url, {'export': 'csv'})
        self.assertEqual(csv_response.status_code, status.HTTP_200_OK)
        self.assertTrue(csv_response['Content-Type'].startswith('text/csv'))
        self.assertIn('ombor_mahsulotlar.csv', csv_response['Content-Disposition'])

    def test_top_kam_qolganlar_and_color_thresholds(self):
        register_url = reverse('register')
        reg_payload = {
            "ism": "TopUser",
            "telefon_raqam": "+998905554433",
            "parol": "123456",
            "parolni_tasdiqlash": "123456",
            "biznes_nomi": "Top Biznes"
        }
        reg_response = self.client.post(register_url, reg_payload, format='json')
        self.assertEqual(reg_response.status_code, status.HTTP_201_CREATED)
        token = reg_response.data['token']
        self.client.credentials(HTTP_AUTHORIZATION=f'Token {token}')

        top_url = reverse('mahsulot-top-kam-qolganlar')
        response = self.client.get(top_url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIsInstance(response.data, list)

    def test_xususiyatlar_catalog_management(self):
        from django.core.cache import cache
        cache.clear()
        register_url = reverse('register')
        reg_payload = {
            "ism": "CatalogUser",
            "telefon_raqam": "+998906667788",
            "parol": "123456",
            "parolni_tasdiqlash": "123456",
            "biznes_nomi": "Catalog Biznes"
        }
        reg_response = self.client.post(register_url, reg_payload, format='json')
        self.assertEqual(reg_response.status_code, status.HTTP_201_CREATED)
        token = reg_response.data['token']
        self.client.credentials(HTTP_AUTHORIZATION=f'Token {token}')

        xususiyatlar_url = reverse('xususiyatlar-list')
        response = self.client.get(xususiyatlar_url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        create_payload = {"nomi": "Rangilari", "tur": "matn"}
        create_res = self.client.post(xususiyatlar_url, create_payload, format='json')
        self.assertEqual(create_res.status_code, status.HTTP_201_CREATED)
        self.assertEqual(create_res.data['nomi'], "Rangilari")

    def test_ombor_filters_and_bulk_action(self):
        from django.core.cache import cache
        cache.clear()
        register_url = reverse('register')
        reg_payload = {
            "ism": "OmborUser",
            "telefon_raqam": "+998907778899",
            "parol": "123456",
            "parolni_tasdiqlash": "123456",
            "biznes_nomi": "Ombor Filter Biznes"
        }
        reg_response = self.client.post(register_url, reg_payload, format='json')
        self.assertEqual(reg_response.status_code, status.HTTP_201_CREATED)
        token = reg_response.data['token']
        self.client.credentials(HTTP_AUTHORIZATION=f'Token {token}')

        list_url = reverse('mahsulot-list')
        res_kritik = self.client.get(list_url, {'holat_rangi': 'kritik'})
        self.assertEqual(res_kritik.status_code, status.HTTP_200_OK)

        bulk_url = reverse('mahsulot-bulk-action')
        bulk_res = self.client.post(bulk_url, {'ids': [1], 'action': 'archive'}, format='json')
        self.assertEqual(bulk_res.status_code, status.HTTP_200_OK)

    def test_xususiyat_soft_delete_and_restore(self):
        from django.core.cache import cache
        cache.clear()
        register_url = reverse('register')
        reg_payload = {
            "ism": "RestoreUser",
            "telefon_raqam": "+998908889900",
            "parol": "123456",
            "parolni_tasdiqlash": "123456",
            "biznes_nomi": "Restore Biznes"
        }
        reg_response = self.client.post(register_url, reg_payload, format='json')
        self.assertEqual(reg_response.status_code, status.HTTP_201_CREATED)
        token = reg_response.data['token']
        self.client.credentials(HTTP_AUTHORIZATION=f'Token {token}')

        xususiyatlar_url = reverse('xususiyatlar-list')
        create_res = self.client.post(xususiyatlar_url, {"nomi": "Artikul", "tur": "matn"}, format='json')
        self.assertEqual(create_res.status_code, status.HTTP_201_CREATED)
        field_id = create_res.data['id']

        detail_url = reverse('xususiyatlar-detail', kwargs={'pk': field_id})
        del_res = self.client.delete(detail_url)
        self.assertEqual(del_res.status_code, status.HTTP_200_OK)

        restore_url = reverse('xususiyatlar-restore', kwargs={'pk': field_id})
        restore_res = self.client.post(restore_url)
        self.assertEqual(restore_res.status_code, status.HTTP_200_OK)

    def test_expanded_ombor_filter_panel(self):
        from django.core.cache import cache
        cache.clear()
        register_url = reverse('register')
        reg_payload = {
            "ism": "FilterUser",
            "telefon_raqam": "+998909990011",
            "parol": "123456",
            "parolni_tasdiqlash": "123456",
            "biznes_nomi": "Filter Panel Biznes"
        }
        reg_response = self.client.post(register_url, reg_payload, format='json')
        self.assertEqual(reg_response.status_code, status.HTTP_201_CREATED)
        token = reg_response.data['token']
        self.client.credentials(HTTP_AUTHORIZATION=f'Token {token}')

        list_url = reverse('mahsulot-list')
        res_holat = self.client.get(list_url, {'holat': 'faol', 'ordering': 'nomi'})
        self.assertEqual(res_holat.status_code, status.HTTP_200_OK)

        res_kat = self.client.get(list_url, {'kategoriya': 'Qurilish'})
        self.assertEqual(res_kat.status_code, status.HTTP_200_OK)

    def test_ombor_sidebar_route_aliases(self):
        from django.core.cache import cache
        cache.clear()
        register_url = reverse('register')
        reg_payload = {
            "ism": "SidebarUser",
            "telefon_raqam": "+998901119988",
            "parol": "123456",
            "parolni_tasdiqlash": "123456",
            "biznes_nomi": "Sidebar Biznes"
        }
        reg_response = self.client.post(register_url, reg_payload, format='json')
        self.assertEqual(reg_response.status_code, status.HTTP_201_CREATED)
        token = reg_response.data['token']
        self.client.credentials(HTTP_AUTHORIZATION=f'Token {token}')

        res_hisobdan = self.client.get(reverse('hisobdan-chiqarish-list'))
        self.assertEqual(res_hisobdan.status_code, status.HTTP_200_OK)

        res_inv = self.client.get(reverse('inventarizatsiya-list'))
        self.assertEqual(res_inv.status_code, status.HTTP_200_OK)

        res_kirim = self.client.get(reverse('kirim-list'))
        self.assertEqual(res_kirim.status_code, status.HTTP_200_OK)




