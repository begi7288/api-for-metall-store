from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError as DRFValidationError
from django.core.exceptions import ValidationError as DjangoValidationError
import django_filters
import openpyxl
from openpyxl.styles import Font, PatternFill
from django.http import HttpResponse

from products.models import Import, XususiyatMaydoni
from products.serializers import ImportSerializer
from user.permissions import IsAdminOrOmborchi
from .common import DynamicPagination, generate_excel_response

class ImportFilter(django_filters.FilterSet):
    sana = django_filters.DateFilter(field_name="yaratilgan_vaqt", lookup_expr='date')

    class Meta:
        model = Import
        fields = ['holat', 'import_turi', 'dokon', 'sana']


class ImportViewSet(viewsets.ModelViewSet):
    serializer_class = ImportSerializer
    permission_classes = [IsAdminOrOmborchi]
    pagination_class = DynamicPagination

    filterset_class = ImportFilter
    search_fields = ['id', 'nomi', 'dokon__nomi']
    ordering_fields = ['miqdori', 'kelish_summasi', 'sotish_summasi', 'yaratilgan_vaqt']

    def list(self, request, *args, **kwargs):
        if request.query_params.get('export') == 'excel':
            queryset = self.filter_queryset(self.get_queryset())
            headers = ["ID", "Nomi", "Do'kon", "Sana", "Jami kelish narxi", "Jami sotish narxi", "Yaratgan xodim"]
            rows = []
            for item in queryset:
                rows.append([
                    item.id,
                    item.nomi,
                    item.dokon.nomi if item.dokon else "",
                    item.sana.strftime("%d.%m.%Y") if item.sana else "",
                    str(item.kelish_summasi),
                    str(item.sotish_summasi),
                    f"{item.yaratgan_xodim.ism} {item.yaratgan_xodim.familiya}" if item.yaratgan_xodim else ""
                ])
            return generate_excel_response("importlar", headers, rows)
        return super().list(request, *args, **kwargs)

    def get_queryset(self):
        user = self.request.user
        queryset = Import.objects.all().order_by('-yaratilgan_vaqt')
        if user.is_superuser:
            return queryset
        if user.is_authenticated and hasattr(user, 'xodim') and user.xodim.biznes:
            return queryset.filter(biznes=user.xodim.biznes)
        return queryset.none()

    def perform_create(self, serializer):
        biznes = None
        executor_xodim = None
        if self.request.user and hasattr(self.request.user, 'xodim'):
            executor_xodim = self.request.user.xodim
            biznes = executor_xodim.biznes
            
        try:
            import_obj = serializer.save(biznes=biznes, yaratgan_xodim=executor_xodim)
        except DjangoValidationError as e:
            if hasattr(e, 'message_dict'):
                raise DRFValidationError(e.message_dict)
            else:
                raise DRFValidationError({'detail': e.messages})
        
        try:
            import_obj.confirm_and_execute(executor_xodim=executor_xodim)
            import_obj.refresh_from_db()
        except Exception as e:
            raise DRFValidationError({'detail': str(e)})

    @action(detail=True, methods=['post'])
    def confirm(self, request, pk=None):
        import_obj = self.get_object()
        executor_xodim = None
        if request.user and hasattr(request.user, 'xodim'):
            executor_xodim = request.user.xodim
            
        try:
            import_obj.confirm_and_execute(executor_xodim=executor_xodim)
        except Exception as e:
            raise DRFValidationError({'detail': str(e)})
            
        return Response({
            'status': "Import muvaffaqiyatli yakunlandi.",
            'holat': import_obj.holat
        }, status=status.HTTP_200_OK)

    def get_permissions(self):
        from rest_framework.permissions import AllowAny
        if self.action == 'template':
            return [AllowAny()]
        return super().get_permissions()

    @action(detail=False, methods=['get'])
    def template(self, request):
        user = request.user
        biznes = None
        if user and user.is_authenticated and hasattr(user, 'xodim'):
            biznes = user.xodim.biznes

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Import Shablon"

        headers = [
            "Nomi", "Shtrix-kod", "Miqdori", "Kelish narxi", "Sotish narxi",
            "Toifa", "Brend", "O'lchov birligi", "Yetkazib beruvchi", "Tavsif"
        ]

        if biznes:
            custom_fields = XususiyatMaydoni.objects.filter(biznes=biznes, is_active=True).order_by('yaratilgan_vaqt')
            for field in custom_fields:
                headers.append(field.nomi)
        else:
            headers.extend(["Qalinligi", "Ranglari"])

        ws.append(headers)

        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 5, 12)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=import_template.xlsx'
        wb.save(response)

        return response
